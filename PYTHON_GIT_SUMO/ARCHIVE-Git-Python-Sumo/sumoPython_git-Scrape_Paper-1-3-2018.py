import pandas as pd
import openpyxl as OPENxlsx
PATH_test_XLSX = '/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES/Network_DF_Period_testFILER.xlsx'

wb = OPENxlsx.Workbook()
# worksheet1 = testXLSX.active
# worksheet1.title = "work sheet number 1"
# worksheet_2 = testXLSX.create_sheet(title = "WS#2")
periodNamesLISTa = list()
for period in range(int(int(estimated_Run_Time)/PERIOD_VARRIABLE)):
    periodNAME = "Period_"+str(((PERIOD_VARRIABLE)*(period))+1)+"_to_"+str((((PERIOD_VARRIABLE)*(period)+1)+PERIOD_VARRIABLE))
    periodNamesLISTa.append(periodNAME)
counter = 0
for i in range(len(periodNamesLISTa[:])):
    ws = testXLSX.create_sheet(title = periodNamesLISTa[i])
testXLSX.save(filename = PATH_test_XLSX)

wb =  OPENxlsx.load_workbook(filename = PATH_test_XLSX)



zzTop = list()
zztop.append('poo thrown all over')
DDFF = pd.DataFrame(zztop)
from openpyxl.utils.dataframe import dataframe_to_rows
for r in dataframe_to_rows(DDFF, index=False, header=True):
    testXLSX[testXLSX.get_sheet_names()[0]].append(r)
    
testXLSX.save(filename = PATH_test_XLSX)



        tabber = "<>"
        print("\n\t\t\t\t<><>Please wait creating Edge Instances<><>\n")
        edgeLISTa = list()
        counter = 0
        for edge_i in Belmont_Ave[0:10]:
            edgeNAME = "edge_"+str(counter) #str(SP.Belmont_AVEDic[counter])
            edgeLISTa.append(edgeNAME)
            # print(edgeLISTa[counter])
            edgeLISTa[counter]= Edge()
            edgeLISTa[counter].set(Belmont_Ave[counter],counter)
            print("Edge = ", counter,tabber,end='\r',flush=True)#, edgeLISTa[counter].__dict__)
            counter = counter + 1
            # tabber = str(tabber+"<>")#\t")
            if (counter/47).is_integer() == True:
                tabber = "\n<>"
            else:
                tabber = str(tabber+"<>")#\t")
        print("\n\t\t\t\t<<<Edge Instances loaded on edgeLISTa>>>\n")
        return edgeLISTa
        
wb.remove_sheet(wb.get_sheet_names()[0])


maxSpeed_i = logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(Belmont_Ave[9]),'Dynamic_Max_Speed']
edge_i = Belmont_Ave[9]

traci.edge.setMaxSpeed(Belmont_Ave[9],maxSpeed_i)