import os, sys
import numpy as np
import traci
import traci.constants as tc
import time
import sumolib
# import General_TraciSUMO_A as GTS
# import Pavement_Condition_A as PC
import cProfile, pstats , io
import pandas as pd
import openpyxl as OPENxlsx
import re

SUMO_outPUT_PREFIX = ""
sumoGUIBinary = "C:/Sumo/SUMO-0.31.0/sumo-0.31.0/bin/sumo-gui"
sumoBinary = "C:/Sumo/SUMO-0.31.0/sumo-0.31.0/bin/sumo"
configPATH = "C:\GitHub\PhD_Modeling\Belmont_AOI_git\Belmont_AOI-runFILES\BMAOI-TRACI.sumocfg"
sumoCmd = [sumoBinary, "-c", configPATH, "--start"]
sumoGUICmd = [sumoGUIBinary, "-c", configPATH, "--start"]
# PATH_Network_DF_Period_0t00_TEMPLATE = '/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/Network_DF_Period_0t00_TEMPLATE.csv'
PATH_Network_DF_Period_0t00_TEMPLATExlsx = '/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES/Network_DF_Period_0t00_TEMPLATE.xlsx'
# PATH_edge_i_cashe_TEMPLATE = '/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/Edge_i_cashe_TEMPLATE.csv' 
PATH_BMAOI_edgeCasheFILES = '/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES'
PATH_Network_DF_Period_0t00_DF = '/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/Network_DF_Period_0t00_DF.csv'

Belmont_Ave=("-12150712#3","-12150712#4","-12150712#6","-12327906#0","-196358954#0","-196358954#3","-196358956#2","-387423966","-423956981","-423956982","-423965484","-423967058#0","-423967058#1","-423967352","-423967353","-423967354","-423967355","-423967356","-423967357","-423967358#1","-423967358#2","-423967359#0","-423967359#1","-424803245","-424803247#1","-424824619","-424824620","-424824621","-424978161","-424978642.0","-424978642.170","-424978643","-424978647#1","-448887867","-448887868","-448887869","-448887870#1","-448887871#0","-448887871#2","-49940170#0","12150712#3","12150712#4","12150712#5","12327906#0","12327906#1","196358954#0","196358954#1","196358956#0","387423966","423956978","423956979","423956980","423965484","423967058#1","423967352","423967353","423967354","423967355","423967356","423967358#0","423967358#2","423967359#0","423967359#1","424803245","424803247#0","424824619","424824620","424824621","424978639.0","424978639.102","424978640","424978643","424978644","424978646","448887867","448887868","448887869","448887870#0","448887871#0","448887871#1")

Belmont_AVEDic ={0: '-12150712#3',1: '-12150712#4',2: '-12150712#6',3: '-12327906#0',4: '-196358954#0',5: '-196358954#3',6: '-196358956#2',7: '-387423966',8: '-423956981',9: '-423956982',10: '-423965484',11: '-423967058#0',12: '-423967058#1',13: '-423967352',14: '-423967353',15: '-423967354',16: '-423967355',17: '-423967356',18: '-423967357',19: '-423967358#1',20: '-423967358#2',21: '-423967359#0',22: '-423967359#1',23: '-424803245',24: '-424803247#1',25: '-424824619',26: '-424824620',27: '-424824621',28: '-424978161',29: '-424978642.0',30: '-424978642.170',31: '-424978643',32: '-424978647#1',33: '-448887867',34: '-448887868',35: '-448887869',36: '-448887870#1',37: '-448887871#0',38: '-448887871#2',39: '-49940170#0',40: '12150712#3',41: '12150712#4',42: '12150712#5',43: '12327906#0',44: '12327906#1',45: '196358954#0',46: '196358954#1',47: '196358956#0',48: '387423966',49: '423956978',50: '423956979',51: '423956980',52: '423965484',53: '423967058#1',54: '423967352',55: '423967353',56: '423967354',57: '423967355',58: '423967356',59: '423967358#0',60: '423967358#2',61: '423967359#0',62: '423967359#1',63: '424803245',64: '424803247#0',65: '424824619',66: '424824620',67: '424824621',68: '424978639.0',69: '424978639.102',70: '424978640',71: '424978643',72: '424978644',73: '424978646',74: '448887867',75: '448887868',76: '448887869',77: '448887870#0',78: '448887871#0',79: '448887871#1'}

equiv_ESAL = {'Car' : 0.0007, 'Truck' : 0.10, 'Semi' : 1.35, '40ftBus' : 1.85}

class Edge():
    import sumolib
    
    # edge_list = []
    def __init__(self):#, edgeID):#, SUMOID):
        # Edge.edge_list.append(self)
        #set came from here
        self.carCount = 0
        self.truckCount = 0
        self.ESAL_TOT = 0
        #self.Original_Max_Speed = 27.78 # (m/s) or 
        
    def set(self, edgeID, Bel_Dic_ID):
        self.edgeID = edgeID
        self.Bel_Dic_ID = Bel_Dic_ID #Belmont_AVEDic[edgeID]
        self.net = sumolib.net.readNet('C:\GitHub\PhD_Modeling\Belmont_AOI_git\Belmont_AOI-runFILES\Belmount_AOI-V5.net.xml')
        self.vehidLIST = {'vehID_k' : 'veh_k_Type'} # https://stackoverflow.com/questions/1024847/add-new-keys-to-a-dictionary
        self.originalMAXSPEED = 27.87# (m/s) self.net.getEdge(edgeID).getSpeed()
    
    def log(self):#, vehID, ESAL_contrib, Emergancy_Stop, Accident):
        Edge_i_VehIDs_lastStep_j = traci.edge.getLastStepVehicleIDs(self.edgeID)
        # if len(Edge_i_VehIDs_lastStep_j) > 0:
            # print("\n",self.Bel_Dic_ID," - Edge_i_VehIDs_lastStep_j = ",Edge_i_VehIDs_lastStep_j)
        for vehID_k in Edge_i_VehIDs_lastStep_j: 
            if vehID_k in self.vehidLIST.keys():
                #print("\nvehID_k is already in here\tvehID_k = ",str(vehID_k))
                continue #?
            else:
                # print("Onward")
                # ##testing vehID_k = Edge_i_VehIDs_lastStep_j[0]
                if re.search('(.*?)Car(.*?)',traci.vehicle.getTypeID(vehID_k),re.IGNORECASE) != None:
                    self.vehidLIST[vehID_k]= traci.vehicle.getTypeID(vehID_k)
                    self.carCount = self.carCount + 1
                    self.ESAL_TOT = self.ESAL_TOT + equiv_ESAL['Car']
                    #print("Car\r\n") #+ edge_i_cashe_csv)
                else:   #Update counts and ESAL Contribution based on type of trcuk ESAL equivalant loads found here http://www.pavementinteractive.org/loads/  
                ##BUT OUTDATED METHOD## New Methods http://www.pavementinteractive.org/equivalent-single-axle-load/
                # https://www.dot.state.pa.us/public/PubsForms/Publications/PUB%20242.pdf - try this out
                    self.vehidLIST[vehID_k]= traci.vehicle.getTypeID(vehID_k)
                    self.truckCount = self.truckCount + 1
                    if re.search('(.*?)Panel(.*)',traci.vehicle.getTypeID(vehID_k),re.IGNORECASE) != None:
                        self.ESAL_TOT = self.ESAL_TOT +  equiv_ESAL['Truck']
                    elif re.search('(.*?)Single_Rear_Truck(.*)',traci.vehicle.getTypeID(vehID_k),re.IGNORECASE) != None:
                        self.ESAL_TOT = self.ESAL_TOT +  equiv_ESAL['Truck']
                    elif re.search('(.*?)Double_Rear_Truck(.*)',traci.vehicle.getTypeID(vehID_k),re.IGNORECASE) != None:
                        self.ESAL_TOT = self.ESAL_TOT +  equiv_ESAL['Semi']
                    elif re.search('(.*?)bus(.*)',traci.vehicle.getTypeID(vehID_k),re.IGNORECASE) != None:
                        self.ESAL_TOT = self.ESAL_TOT +  equiv_ESAL['40ftBus']
                    else:
                        print("I Don't Know What you Are")
                        
    @classmethod # means class instead of self (Edge.create_Edge_Instances() becomes create_Edge_Instances(Edge))
    def create_Edge_Instances(cls):
        # @staticmethod # means no self
        # def create_Edge_Instances():
        tabber = "<>"
        print("\n\t\t\t\t<><>Please wait creating Edge Instances<><>\n")
        edgeLISTa = list()
        counter = 0
        for edge_i in Belmont_Ave[:]:
            edgeNAME = "edge_"+str(counter) #str(SP.Belmont_AVEDic[counter])
            edgeLISTa.append(edgeNAME)
            # print(edgeLISTa[counter])
            edgeLISTa[counter]= cls()
            edgeLISTa[counter].set(Belmont_Ave[counter],counter)
            print("Edge = ", counter,tabber,end='\r',flush=True)#, edgeLISTa[counter].__dict__)
            counter = counter + 1
            # tabber = str(tabber+"<>")#\t")
            if (counter/47).is_integer() == True:
                tabber = "\n<>"
            else:
                tabber = str(tabber+"<>")#\t")
        print("\n\t\t\t\t<<<Edge Instances loaded on edgeLISTa>>>\n")
        return edgeLISTa
        
    def __str__(self): #[str(edge) for edge in edgeLISTa]
        return str(self.__dict__)
    
    def __repr__(self):
        return repr(self.__dict__)
        
    # def edgeLIST(self):
        # for edge_i in SP.Belmont_Ave[:]:
            # edgeNAME = "edge_"+str(SP.Belmont_AVEDic[edge_i])
        # print(edgeNAME," - ", str(edge_i))
        # setattr(self, edgeNAME, SP.Edge(str(edge_i)))
        
        
    def _restLOGGER(self,edgeLISTa):
        print("\n\nReset_Edge_CasheFiles has started")
        counter = 0
        for edge_i in Belmont_Ave[:]:
            edgeLISTa[counter].vehidLIST= ()
            counter = counter + 1
        return print("Reset_Edge_CasheFiles has completed\n\n")

    def setNewMaxSpeed(self,logger_TEMPDF):
        print("\n\n<><><><Changing SUMO EGDE MAX SPEED><><><>\n")
        for rd in Belmont_Ave:
            maxSpeed_i = logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(rd),'Dynamic_Max_Speed']
            maxSpeed_i = round(maxSpeed_i.iloc[0],2)
            traci.edge.setMaxSpeed(rd,maxSpeed_i)
            if maxSpeed_i <= 27:
                print("The new speed for edge ",rd," is now ",round(maxSpeed_i.iloc[0],2))
        
        
       
class Network_Period:
    def load_n_create_Excel_NetworkFile(SUMO_outPUT_PREFIX, PERIOD_VARRIABLE,steps_TT, display = None, PATH=None):
        if display != 0:
            print("Loading and creating Excel Network File", "SUMO_outPUT_PREFIX = ", SUMO_outPUT_PREFIX)
        if PATH == None:
            PATH_Network_DF_Period_0t00_TEMPLATExlsx = '/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES/Network_DF_Period_0t00_TEMPLATE.xlsx'
            PATH_to_Save_to = "/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES/Network_" +SUMO_outPUT_PREFIX + "_PeriodBook.xlsx"
        wb = OPENxlsx.Workbook()
        # worksheet1 = testXLSX.active
        # worksheet1.title = "work sheet number 1"
        # worksheet_2 = testXLSX.create_sheet(title = "WS#2")

        estimated_Run_Time = 90000 #steps_TT
        Network_DF_Period_0t00xlsx=pd.read_excel(PATH_Network_DF_Period_0t00_TEMPLATExlsx)
        periodNamesLISTa = list()
        for period in range(int(int(estimated_Run_Time)/PERIOD_VARRIABLE)):
            periodNAME = "Period_"+str(((PERIOD_VARRIABLE)*(period))+1)+"_to_"+str((((PERIOD_VARRIABLE)*(period)+1)+PERIOD_VARRIABLE))
            periodNamesLISTa.append(periodNAME)
        # print("/n<><><periodNamesLISTa = ",periodNamesLISTa)
        from openpyxl.utils.dataframe import dataframe_to_rows
        for i in range(0,len(periodNamesLISTa[:])):
            # print("\nwb.get_sheet_names()[0] = ",wb.get_sheet_names()[0])
            ws = wb.create_sheet(title = periodNamesLISTa[i])
        if wb.get_sheet_names()[0] == 'Sheet':
            print("\nwb.get_sheet_names()[0] = ",wb.get_sheet_names()[0])
            delme = wb.get_sheet_by_name('Sheet')
            print("\nRemoving wb.get_sheet_names()[0] = ",wb.get_sheet_names()[0])
            wb.remove_sheet(delme)
            wb.save(filename = PATH_to_Save_to)
            # for r in dataframe_to_rows(Network_DF_Period_0t00xlsx, index=False, header=True):
                # wb[wb.get_sheet_names()[i]].append(r)
        
        wb.save(filename = PATH_to_Save_to)
        wb.close()
        return wb, periodNamesLISTa 
        
    def fillOutworksheet(SUMO_outPUT_PREFIX,periodCounter,edgeLISTa,periodNamesLISTa):
        #https://chrisalbon.com/python/data_wrangling/pandas_dataframe_load_xls/
        if periodCounter == 0:
            return
        # from openpyxl import load_workbook
        PATH_Network_DF_Period_0t00_TEMPLATExlsx = '/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES/Network_DF_Period_0t00_TEMPLATE.xlsx'

        print("\n\n<><><periodCounter = ",int(round(periodCounter)))#,">>>\wb[wb.get_sheet_names()[periodCounter+1]] = ",wb[wb.get_sheet_names()[periodCounter+1]])#, " ...But...\nNow periodCounter = ", round(periodCounter))
        periodCounter = int(round(periodCounter))#periodCounter = int(round(periodCounter)-1) # "-1" because we want to fill in the sheet from the perivious period.
        logger_TEMPDF =pd.read_excel(PATH_Network_DF_Period_0t00_TEMPLATExlsx) #grabs the currennt period sheet
        print(type(logger_TEMPDF))
        logger_TEMPDF = pd.DataFrame(logger_TEMPDF)
        print(type(logger_TEMPDF))
        counter = 0
        for edge_i in Belmont_Ave[:]: 
            logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edge_i),'Total_Vehicles'] = (edgeLISTa[counter].truckCount + edgeLISTa[counter].carCount)
            logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edge_i),'Total_Trucks'] = edgeLISTa[counter].truckCount
            logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edge_i),'ESAL'] = edgeLISTa[counter].ESAL_TOT
            Condition_RTi = 100.00 - (edgeLISTa[counter].ESAL_TOT) * 0.01339
            logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edge_i),'Condition_Index'] = Condition_RTi
            # maxSpeedo = logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edge_i),'Original_Max_Speed']
            maxSpeedo = edgeLISTa[counter].originalMAXSPEED
            maxSpeed_i =(maxSpeedo - (maxSpeedo**((100-Condition_RTi)/96))+4.5675) #Units m/s
            logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edge_i),'Dynamic_Max_Speed'] = maxSpeed_i
            if counter == 4:
                print("edgeLISTa[counter].truckCount = ",edgeLISTa[counter].truckCount,"<><>\nlogger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edge_i),'Total_Trucks']= ",logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edge_i),'Total_Trucks'],"<><>\n")
            counter += 1
        ### NOW I WANT TO SAVE THIS DATAFRAME [[CURRENT_SHEET]] TO THE SHEET THAT IT CAME FROM IN THE WORKBOOKER
        # print("\n logger_TEMPDF = \n",logger_TEMPDF, "\n<><>\n")
        # from openpyxl.utils.dataframe import dataframe_to_rows  ##http://openpyxl.readthedocs.io/en/default/tutorial.html
        # http://openpyxl.readthedocs.io/en/default/pandas.html#working-with-pandas-dataframes
        ### TRY THIS http://pbpython.com/improve-pandas-excel-output.html and then this https://codereview.stackexchange.com/questions/180405/writing-excelsheet-from-python-dataframe
        # for r in dataframe_to_rows(logger_TEMPDF, index=False, header=True):
            # print(r)
            # wb[wb.get_sheet_names()[periodCounter-1]].append(r) # "-1" because we want to fill in the sheet from the perivious period.
        # logger_TEMPDF.to_excel(writer, sheet_name =periodNamesLISTa[periodCounter])# 
        Network_Period.myWrite_to_excel(logger_TEMPDF,periodCounter,SUMO_outPUT_PREFIX)
        print("\n>>>Data written to:: wb[wb.get_sheet_names()[periodCounter]] = ",wb[wb.get_sheet_names()[periodCounter-1]],"\n")
        # wb.close()
        Edge.setNewMaxSpeed(edgeLISTa,logger_TEMPDF)
        if periodCounter != 1:
            print("edgeLISTa[4] = ",edgeLISTa[4])
        ## Why do I need to reset the logger? Will it slow things down towards the end? 
        ## Edge._restLOGGER(edgeLISTa)
        return wb 
            
    def myWrite_to_excel(logger_TEMPDF,periodCounter,SUMO_outPUT_PREFIX):
        ## openpyxl.worksheet.Worksheet.cell() method from  https://openpyxl.readthedocs.io/en/default/tutorial.html#data-storage
        PATH_to_Save_to = "/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES/Network_" +SUMO_outPUT_PREFIX + "_PeriodBook.xlsx"
        wb =  OPENxlsx.load_workbook(filename = PATH_to_Save_to)
        periodCounter = int(round(periodCounter))
        ws = wb[wb.get_sheet_names()[periodCounter-1]]
        logger_TEMPDF =pd.read_excel(PATH_Network_DF_Period_0t00_TEMPLATExlsx)
        counter_i = 0
        counter_j = 0
        for i in range(logger_TEMPDF.shape[0]):
            counter_i +=1
            counter_j = 0
            for j in range(logger_TEMPDF.shape[1]):
                counter_j +=1
                if i == 0:
                    ws.cell(row=counter_i, column=counter_j, value=logger_TEMPDF.columns[j])
                    ws.cell(row=counter_i+1, column=counter_j, value=logger_TEMPDF.iloc[i,j])
                else:
                    ws.cell(row=counter_i+1, column=counter_j, value=logger_TEMPDF.iloc[i,j])
        wb.save(PATH_to_Save_to)
        return wb
        
class Initializer:
    
    def __init__(self,new):
        return #new
        
        
    def inputPeriod_asNumber(new):
        #global PERIOD_VARRIABLE
        if new ==1:
            while True: #http://www.101computing.net/number-only/
                    try:
                        userINPUT = input("\r\n\nPress [1] to enter a user defined Period_Interval.\r\nPress anything else to continue default is seconds 3600 or 1 hour? ... ") 
                        if  userINPUT != '1':
                            PERIOD_VARRIABLE = 3600 #Default to 1 hour
                            print("\n\nPERIOD_VARRIABLE = ", PERIOD_VARRIABLE,"\n")
                            return int(PERIOD_VARRIABLE)
                            break
                        else:
                            PERIOD_VARRIABLE = int(input("\nWhat would you like the Period of Collection to be default is seconds 3600 or 1 hour? ... "))
                        # while PERIOD_VARRIABLE.is_integer == False:
                            # PERIOD_VARRIABLE = input("!!!!!!!!!!!!Type ERROR\t\tPlease Enter an Integer\n\t\t\t\tWhat would you like the Period of Collection to be default is seconds 3600 or 1 hour? ... ")
                    except ValueError:# as err: #https://docs.python.org/3/library/exceptions.html#BaseException 
                        print("\n!!!!!!!!!!!!Type ERROR\n\t\tPlease Enter an Integer\n\tWhat would you like the Period of Collection to be default is seconds 3600 or 1 hour? ... ")
                        continue
                    else:
                        print("PERIOD_VARRIABLE = ", PERIOD_VARRIABLE)
                        return int(PERIOD_VARRIABLE)
                        break
        else:
            print("PERIOD_VARRIABLE = ", PERIOD_VARRIABLE)
            return PERIOD_VARRIABLE


    def startSUMO(SUMO_Traci_PORT,useCase=None):
        sumoCmd = [sumoBinary, "-c", configPATH, "--start"]
        sumoGUICmd = [sumoGUIBinary, "-c", configPATH, "--start"]
        portTouse = int(SUMO_Traci_PORT)
        # global Start_Time
        # while rideAgain == 1:
            #print("useCase =",useCase)
        if useCase == None:
           print("\t\t\t\t\t\t###STARTING####")
           useCase = input("\n\tPress 1 for a new Run:  \n") #\nPress Anything to continue...
        if useCase == "1":
            GUI_01 = input("Press Zero to use GUI? (0)\n")
            inputPort = input("Confirm with: BMAOI-TRACI.sumocfg .. port in File: .."+ str(SUMO_Traci_PORT)+"\nPlease select a port\t...") # 
            if inputPort == '':
                portTouse = int(SUMO_Traci_PORT)
            else:
                portTouse = inputPort
            print("\n\t\tTrying port ..  ",portTouse,"\n\n")
            if str(GUI_01) != "0":
                traci.start(sumoCmd,int(portTouse))
            else:
                traci.start(sumoGUICmd,int(portTouse))
            traci.simulationStep()
            Start_Time = int(traci.simulation.getCurrentTime()/1000)
            #useCase == "Continue"
        elif useCase == "Continue":
            print("\t\t\t\t   !!!!!!!!!!!!!!CONTINUING!!!!!!!!!!!!!\t\t\t")
            typeRun = "4"
            Start_Time = int(traci.simulation.getCurrentTime()/1000)
        elif useCase == "x":
            PC.condition.DisplayFiles()
            traci.close()
        print("useCase = ",useCase)
        return Start_Time, portTouse, useCase

        
class Runner:
    # def __init__(PERIOD_VARRIABLE,periodCounter=None):
        # self.PERIOD_VARRIABLE = PERIOD_VARRIABLE
        # self.periodDIC = {}
        # if periodCounter == None:
            # self.periodDIC={'period_1.0' : Network_Period(self.PERIOD_VARRIABLE,SUMO_outPUT_PREFIX)}
        # else:
            # self.periodDIC={"period_"+str(periodCounter):Network_Period(PERIOD_VARRIABLE,SUMO_outPUT_PREFIX)}

    def runtypeAsker(typeRun=None):
        print("\nBeginning Journey Now\n\tPlease See Gui for Show\n")
        typeRun = input("\n\nPush T to Run Simulation until the end\nPush 2 to run for XXX Steps\nPush 3 to run until set truck accomulation level\nPush 4 to run until input time\nCurrent Step_Time : "+str(traci.simulation.getCurrentTime()/1000)+"\n\n\tEnter Run Type Here: ")
        if typeRun == '':
            typeRun = "2"
            print("You didn't specify! Going with default value... typeRun =", str(typeRun))
        else:
            print("You pressed", str(typeRun),"= typeRun")
        return typeRun
    
        ###Run the Code####
    def releaseTraci(Start_Time,typeRun,edgeLISTa,PERIOD_VARRIABLE,SUMO_outPUT_PREFIX,periodNamesLISTa,steps_TT=None):
        #Defining the Constants
        if steps_TT == None:
            steps_TT = 0
        stepCounter = 0
        addTesterTrucks = 0
        #Initializer.runSUMO(SUMO_Traci_PORT)[0]
        Period_Break_Point=('50','1000','2500')
        periodCounter = (traci.simulation.getCurrentTime()/1000)#/PERIOD_VARRIABLE # was 0
        Sim_Step_Count = 0
        # if typeRun == None:
            # typeRun = 2
        if str(typeRun) == "T": #Run Simulation unti the end
            while traci.simulation.getMinExpectedNumber() > 0:
                #print("step No. = ", step," out of ",No_next_steps) #Showing the steps boggy downy machiny
                traci.simulationStep()
                Runner.thingsTodoWhileStepping(edgeLISTa,PERIOD_VARRIABLE,periodCounter,SUMO_outPUT_PREFIX,periodNamesLISTa)
                periodCounter += 1
        elif str(typeRun) == "2": #run for XXX Steps VAR -> steps_TT
            #print("\nCheck Check... typeRun is: ",typeRun,"=2(?)")
            #addTesterTrucks = input("\nDO you want to add some test truckers (1/0)?")
            estimated_Run_Time = input("...\n\n\nHow long will you run this file for... ")
            if estimated_Run_Time == '':
                estimated_Run_Time = 90000
                print("You did not specify how long you wanted to run until so the default value = ", estimated_Run_Time)
                steps_TT = int(estimated_Run_Time)
            if steps_TT == None:
                steps_TT = input("How many steps would you like to take? ")
            #No_next_steps = int(steps_TT)
            if addTesterTrucks == "1":
                traci.simulationStep()
                general.addBen_Trackers()
            for step in range(int(steps_TT)):
                traci.simulationStep()
                Runner.thingsTodoWhileStepping(edgeLISTa,PERIOD_VARRIABLE,periodCounter,SUMO_outPUT_PREFIX,periodNamesLISTa)
                periodCounter += 1#(1/PERIOD_VARRIABLE)
                ###PC.condition.Truck_Count_III(useCase, Sim_Step_Count, PERIOD_VARRIABLE = int(PERIOD_VARRIABLE))
        elif str(typeRun) == "3": # run until set truck accomulation level Var -> TotTruckLevel
            while int(general.Edge_TOT_Trucks) < 600: #runs until you find the 1st (max) total trucks that equal a number, what if it was the last (min)
                #print("step No. = ", step," out of ",No_next_steps) #Showing the steps boggy downy machiny
                traci.simulationStep()
                ### Make this a function ####
                PC.condition.Truck_Count_III(useCase, Sim_Step_Count, PERIOD_VARRIABLE = int(PERIOD_VARRIABLE))
                steps_TT = steps_TT +1
        elif str(typeRun) == "4": #run until input time VAR -> NextTimeToStop			
            print("\nCurrent Time is: ",str(traci.simulation.getCurrentTime()/1000))
            # import ipdb
            # ipdb.set_trace()
            NextTimeToStop = input("\nWhen would you like to run until? ... ")
            NextTimeToStop = int(NextTimeToStop)
            #######PERIOD_VARRIABLE = input("\nWhat would you like the Period of Collection to be default is seconds 28801 or 8:00am? ... ")
            print("\t\t\t<><>Running until time: ",NextTimeToStop,"<><>")
            while traci.simulation.getCurrentTime()/1000 < NextTimeToStop:
                ##WHAT I NEED HERE IS A WAY TO GENERATE NEW NETWORK_DFs PER PERIOD
                traci.simulationStep()
                Runner.thingsTodoWhileStepping(edgeLISTa,PERIOD_VARRIABLE,periodCounter,SUMO_outPUT_PREFIX,periodNamesLISTa)
                periodCounter += 1#(1/PERIOD_VARRIABLE)
                # if (steps_TT/10).is_integer:
                    # print(" steps_TT = ",steps_TT, "Sim_Step_Count = ",Sim_Step_Count)
        #aperiodDIC = pp.thingsTodoWhileStepping(edgeLISTa,PERIOD_VARRIABLE,periodCounter,SUMO_outPUT_PREFIX) #Runner became self
        print("<>\n<><>\n<><><>\nperiodCounter = ",periodCounter)#, "periodNamesLISTa = ",periodNamesLISTa)
        return # edgeLISTa
                    
    def thingsTodoWhileStepping(edgeLISTa,PERIOD_VARRIABLE,periodCounter,SUMO_outPUT_PREFIX,periodNamesLISTa):
        for i in range(len(Belmont_Ave)):
            edgeLISTa[i].log()
        ##Every minute ie 60 seconds update Network file
        ##Every Period length save Network as a new sheet in a workbook
        # if round(periodCounter,len(str(PERIOD_VARRIABLE))+1) %1 == 0:
            # print("\n\n\n<><> THIS IS A TEST <><><><><>\n len(str(PERIOD_VARRIABLE)+1)= ",len(str(PERIOD_VARRIABLE)),"\n<><><><><><periodCounter><><><><><>",periodCounter,"\nAAAAnd round(periodCounter,len(str(PERIOD_VARRIABLE))+1) %1 = ",(round(periodCounter,len(str(PERIOD_VARRIABLE))+1) %1))
            # if round(periodCounter,len(str(PERIOD_VARRIABLE))+1) %1 == 0: ## if should still go here
                # periodCounter = round(periodCounter,0)
                # Network_Period.fillOutworksheet(SUMO_outPUT_PREFIX,periodCounter,edgeLISTa,periodNamesLISTa,workBooker)
        if (periodCounter/PERIOD_VARRIABLE).is_integer():
            periodCounter = (periodCounter/PERIOD_VARRIABLE)
            Network_Period.fillOutworksheet(SUMO_outPUT_PREFIX,periodCounter,edgeLISTa,periodNamesLISTa)
            print("\n\n\n<><>edgeLISTa[4] = ", edgeLISTa[4])
        # return edgeLISTa
                    

    def other_parts_not_yet_implemented():
        ## General Output and things after each run has done its things
        #Show Network files
        #print("Network_DF_Period_PATH = ",Network_DF_Period_PATH, "Network_DF_Period_DF = ", Network_DF_Period_DF)
        PC.condition.DisplayFiles(PERIOD_VARRIABLE)
        # print("\n\nThanks for riding the SUMO Traci Train. Please come back again soon\r\n$ run /GitHub/PhD_Modeling/PYTHON_GIT_SUMO/SUMO_Runner.py\n")
        # goAgain = input("\n\n >>>>>>> This part of the journey has ended. Press anything to contine?\n>>>>>>>>>Or x to EXIT\n\n>  >\t> >\t>> ")
        # if goAgain == "x":
# #            traci.close()
            # print("\n\n^^---Journey has ended---^^\n\n") #,edge_VehIDhistoryPD,"
        # else:
            # #general.addNewColumn(2,Period_Time)
            # general.releaseTraci(useCase="Continue")
        general.outPUT(useCase,Start_Time,PERIOD_VARRIABLE, steps_TT)
        
class RunFileInfo:
    


    def __init__(self):#, SUMO_outPUT_PREFIX,SUMO_Traci_PORT):
        pass
        
    def GetSimulationRunPrefix(display):
        import re
        global SUMO_outPUT_PREFIX
        global SUMO_Traci_PORT
        with open('/GitHub/PhD_Modeling/Belmont_AOI_git/Belmont_AOI-runFILES/BMAOI-TRACI.sumocfg') as f:
            for line in f:
                if '<output-prefix value="' in line:
                    print(line)
                    SUMO_outPUT_PREFIX_LINE = line
                    SUMO_outPUT_PREFIX = re.search('<output-prefix value="(.*)-(.*)-"/>',SUMO_outPUT_PREFIX_LINE, re.IGNORECASE).group(2)
                    # print("\nSUMO_outPUT_PREFIX = ",SUMO_outPUT_PREFIX,"\n")
                    break
            for line in f:
                if '<remote-port value=' in line:
                    SUMO_outPUT_Port_LINE = line
                    SUMO_Traci_PORT = re.search('<remote-port value="(.*)"/>',SUMO_outPUT_Port_LINE, re.IGNORECASE).group(1)
        if display == 1:
            print("\nSUMO_outPUT_PREFIX = ", SUMO_outPUT_PREFIX,"\n\SUMO_Traci_PORT = ", SUMO_Traci_PORT)

        return SUMO_outPUT_PREFIX, SUMO_Traci_PORT#, print("\nSUMO_outPUT_PREFIX = ", SUMO_outPUT_PREFIX,"\nSUMO_Traci_PORT = ", SUMO_Traci_PORT)
        
print("Your new sumoPython_git_A has been loaded!")
        ## Python Sumo Script

## Grave yard
    # def thingsTodoWhileStepping(edgeLISTa,PERIOD_VARRIABLE,periodCounter,SUMO_outPUT_PREFIX):
        # # self.periodDIC={"period_"+str(periodCounter):Network_Period(PERIOD_VARRIABLE,SUMO_outPUT_PREFIX)}
        # # print(type(periodCounter), "periodCounter = ", periodCounter, "periodLISTa = ",periodLISTa)
        # if periodCounter == 0:
            # return
        # elif round(periodCounter,4) == 1:
            # print(type(round(periodCounter,4)), "periodCounter = ", periodCounter, "periodLISTa = ",periodLISTa)
            # self.periodDIC["period_"+str(round(periodCounter,0))].excelSAVER()
            # # self.periodDIC["period_"+str(round(periodCounter,0)+1)] = Runner(PERIOD_VARRIABLE,round(periodCounter,4))#Network_Period(PERIOD_VARRIABLE,SUMO_outPUT_PREFIX)
            # self.periodDIC["period_"+str(round(periodCounter,0)+1)] = Network_Period(PERIOD_VARRIABLE,SUMO_outPUT_PREFIX,round(periodCounter,4))
        # elif round(periodCounter,4) %1 == 0:#.is_integer():#test to see if it is a whole number which would require a new periodFILE  (traci.simulation.getCurrentTime()/1000)/PERIOD_VARRIABLE
            # print(type(round(periodCounter,4)), "periodCounter = ", periodCounter, "periodLISTa = ",periodLISTa)
            # self.periodDIC["period_"+str(round(periodCounter,0))].excelSAVER()
            # self.periodDIC={"period_"+str(round(periodCounter,0)):Network_Period(PERIOD_VARRIABLE,SUMO_outPUT_PREFIX)}
            # # self.periodDIC["period_"+str(round(periodCounter,0)+1)] = Runner(PERIOD_VARRIABLE,round(periodCounter,4))
            # self.periodDIC["period_"+str(round(periodCounter,0)+1)] = Network_Period(PERIOD_VARRIABLE,SUMO_outPUT_PREFIX,round(periodCounter,4))
            # # self.periodDIC={"period_"+str(round(periodCounter,0)):Network_Period(PERIOD_VARRIABLE,SUMO_outPUT_PREFIX)} #self.periodDIC
            # # periodDIC_to_save={}
            # # periodDIC_to_save = periodDIC #self.periodDIC
            # # periodDIC["period_"+str(periodCounter)].excelSAVER()
            # print("\n\n....\t...\t..\t.\t<<!>><<>>OLD Period<<>><<!>>\n periodFILE = ",self.periodDIC["period_"+str(round(periodCounter,4))].currentNetwork_DF_Period_PATH,"\n",self.periodDIC["period_"+str(round(periodCounter,4))].currentNetwork_DF_FILE.iloc[38],"\n","Sheetname = ",self.periodDIC["period_"+str(round(periodCounter,4))].currentPeriodSheetname,"\n")
            # startPeriodTime = int(traci.simulation.getCurrentTime()/1000)
            # endPeriodTime = startPeriodTime + PERIOD_VARRIABLE
            # #Network_Period.excelSAVER
            # # legacy condition.GetSimulationRunPrefix(display = 0)
            # print("startPeriodTime = ",startPeriodTime, "endPeriodTime = ",endPeriodTime)
