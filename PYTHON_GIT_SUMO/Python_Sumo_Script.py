# Ask for Parameters
    # Period Length
# import os, sys
import numpy as np
import traci
import traci.constants as tc
import re
import time
import sumolib
# import cProfile, pstats , io
import pandas as pd
import sumoPython_git_A as SP
import openpyxl as OPENxlsx
# %colors Linux
PERIOD_VARRIABLE = SP.Initializer.inputPeriod_asNumber(new=1)
fileINFO = SP.RunFileInfo.GetSimulationRunPrefix(display=0)
SUMO_outPUT_PREFIX = SP.RunFileInfo.GetSimulationRunPrefix(display=0)[0]
SUMO_Traci_PORT = int(SP.RunFileInfo.GetSimulationRunPrefix(display=0)[1])
SP.Initializer.startSUMO(SUMO_Traci_PORT,useCase=str(1))
    # Ask for Steps to take or Time to run until
typeRun = SP.Runner.runtypeAsker()
Start_Time = int(traci.simulation.getCurrentTime()/1000) #SP.Initializer.runSUMO(SUMO_Traci_PORT,useCase="Continue")[0]
# estimated_Run_Time = input("...\n\n\nHow long will you run this file for... ")
# if estimated_Run_Time == '':
estimated_Run_Time = 90000
    # print("You did not specify how long you wanted to run until so the default value = ", estimated_Run_Time)
steps_TT = int(estimated_Run_Time)
# Initalize Files
edgeLISTa = SP.Edge.create_Edge_Instances()
wb = SP.Network_Period.load_n_create_Excel_NetworkFile(SUMO_outPUT_PREFIX,PERIOD_VARRIABLE,steps_TT,PATH=None)[0]
periodNamesLISTa = SP.Network_Period.load_n_create_Excel_NetworkFile(SUMO_outPUT_PREFIX,PERIOD_VARRIABLE,steps_TT,display=0)[1]

# Take a step(s)
SP.Runner.releaseTraci(Start_Time,typeRun,edgeLISTa,PERIOD_VARRIABLE,SUMO_outPUT_PREFIX,periodNamesLISTa,steps_TT)


### I am up to here ###





# ##workin on saving excel correctly
# import openpyxl as OPENxlsx
# workBooker = OPENxlsx.Workbook()
# PATH_to_Save_to = "/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES/Network_" +SUMO_outPUT_PREFIX + "_PeriodBook.xlsx"
# workBooker = load_workbook(PATH_to_Save_to)
# workBooker.create_sheet(title=periodNamesLISTa[periodCounter])





    # call each edge
        # get vehID_List
    # call logger
        # compare vehID_List with logger

        # Add new vehIDs to loggers
            # compute stats for vehID
    # Every minute update Network Period Files
        # Update Sumo Edge Parameters
        
        
        
# edge_i_cashe_csv = pd.read_csv(PATH_edge_i_cashe_TEMPLATE)
# edge_cashe_LOAD_PATH = "/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES" + "/" + PC.condition.Belmont_AVEDic[edge_i] +".csv"

###Run until we have values
# while bool(Edge_i_VehIDs_lastStep_j) is False:
     # traci.simulationStep()
     # Runner.thingsTodoWhileStepping(edgeLISTa,PERIOD_VARRIABLE,periodCounter,SUMO_outPUT_PREFIX,periodNamesLISTa)
     # Edge_i_VehIDs_lastStep_j = traci.edge.getLastStepVehicleIDs(edgeLISTa[4].edgeID)
     
     

    
# Try a few things:
# change up dictionary
# hard code
# add hourly cars to evaluate network performance


### Changing up Template 1-8-18 ##

# PATH_Network_DF_Period_0t00_TEMPLATExlsx = '/GitHub/PhD_Modeling/Belmont_AOI_git/Belmont_AOI-runFILES/Network_DF_Period_0t00_TEMPLATE.xlsx'
# newTemplate = pd.read_excel(PATH_Network_DF_Period_0t00_TEMPLATExlsx)
# for i in range(len(edgeLISTa)):
    # newTemplate.loc[i,'Belmont_AVEDic_ID'] = edgeLISTa[i].edgeID
# pd.DateFrame(newTemplate)
# # newTemplate.to_excel(PATH_Network_DF_Period_0t00_TEMPLATExlsx)
# for i in range(75,len(edgeLISTa)):
    # if Belmont_AVEDic[i] == edgeLISTa[i].edgeID == logger_TEMPDF.loc[i,'Belmont_AVEDic_ID']:
        # print(" Matches ",edgeLISTa[i].edgeID," Hell's Yeah!")
        
# edgeLISTa[n].truckCount == logger_TEMPDF.loc[n,'Total_Trucks']

# logger_TEMPDF.loc[n,'Total_Trucks'] == edgeLISTa[n].truckCount
# wsdf.loc[n,7] == edgeLISTa[n].truckCount


# edges = {SP.Belmont_AVEDic[edge]: SP.Edge() for edge in SP.Belmont_Ave}
    # %recall = SP.Edge(str(edge_i))
    # setattr(self, edgeNAME, SP.Edge(str(edge_i)))
    # # exec(edgeNAME + " = SP.Edge(str(edge_i))")
    # # str("edge_"+str(SP.Belmont_AVEDic[edge_i])) = SP.Edge(str(edge_i))
# # edge_73e = SP.Edge('424978646') or now with the set function
# edge_73 = SP.Edge()
# edge_73.set('424978646',73)
    # # for edge_i in SP.Belmont_Ave[:]:
        # # edgeNAME = "edge_"+str(SP.Belmont_AVEDic[edge_i])
        # # print(edgeNAME," - ", str(edge_i))
        # # setattr(SP.Edge,edgeNAME,str(edge_i))
        
# for i in range(10):
    # tabber = "a\t"
    # print(tabber,"a",end='\r',flush=True)
    # tabber = str(tabber+"\t")

## How to tell if periodCounter is a whole number    
# import time
# periodCounter = 0
# PERIOD_VARRIABLE=1800
# for i in range(3600):
# periodCounter += (1/PERIOD_VARRIABLE)
# if round(periodCounter,4).is_integer():# == int(periodCounter):  #% 1 == 0:
    # whole_numb = round(periodCounter,4)
    # print("Whole Number?", whole_numb,"\n")
# else:
    # time.sleep(0.005)
        # print("periodCounter = ",(periodCounter))#,end='\r',flush=True)