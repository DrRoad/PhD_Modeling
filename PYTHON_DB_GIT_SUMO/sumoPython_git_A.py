## Last worked on 2/22/2018
# Primary Code for PhD Hybrid ABM Transportation Asset Management Planning
# Handles Starting SUMO 
# and Collecting Metrics

import os, sys
import numpy as np
import traci
import traci.constants as tc
import time
import sumolib
# import General_TraciSUMO_A as GTS
# import Pavement_Condition_A as PC
import cProfile, pstats , io
import pandas as pd
import openpyxl as OPENxlsx
import re
import psutil #https://stackoverflow.com/questions/276052/how-to-get-current-cpu-and-ram-usage-in-python
memoryUse = 0
pid = os.getpid()
py = psutil.Process(pid)
##### periodCounter = 0#/PERIOD_VARRIABLE # was 0
sumoGUIBinary = "C:/Sumo/sumo-0.32.0/bin/sumo-gui-0.32.0"
sumoBinary = "C:/Sumo/sumo-0.32.0/bin/sumo"


# sumoGUIBinary = "C:/Sumo/bin/sumo-gui"
# sumoBinary = "C:/Sumo/bin/sumo"
# sumoGUIBinary = "C:/Sumo/SUMO-0.31.0/sumo-0.31.0/bin/sumo-gui"
# sumoBinary = "C:/Sumo/SUMO-0.31.0/sumo-0.31.0/bin/sumo"
configPATH = "C:/Dropbox/Phd_R_Ms/PhD_Modeling_DB_GIT/Belmont_AOI_git/Belmont_AOI-runFILES/BMAOI-TRACI-DB.sumocfg"
sumoCmd = [sumoBinary, "-c", configPATH, "--start"]
sumoGUICmd = [sumoGUIBinary, "-c", configPATH, "--start"]


SUMO_outPUT_PREFIX = ""#"BMAO_RUN_4_V30"
# sumoCmd = [sumoBinary,"--remote-port",SUMO_Traci_PORT,"--begin 0 --step-length 1 --net-file C:\Dropbox\Phd_R_Ms\PhD_Modeling_DB_GIT\Belmont_AOI_git\Belmont_AOI-runFILES\Belmount_AOI-V5.net.xml --additional-files ...--vehroute-output.exit-times true --vehroute-output.sorted true --vehroute-output.intended-depart true --vehroute-output.route-length true --vehroute-output.write-unfinished false --load-state true --no-warnings false --no-step-log false --duration-log.statistics true --ignore-route-errors false --step-method.ballistic false --collision.action teleport --collision.stoptime 20 --collision.check-junctions false --waiting-time-memory 10 --time-to-impatience 120 --max-depart-delay -1 --device.rerouting.probability 0.85 --device.rerouting.period 60 --device.rerouting.output C:\Sumo\runs\BelmontC_AOI_main\BelmontC_AOI-outPUT\BMAOI_C-VehRouteFILES\RE-Routing_output.xml --start"]

# PATH_Network_DF_Period_0t00_TEMPLATE = '/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/Network_DF_Period_0t00_TEMPLATE.csv'
PATH_Network_DF_Period_0t00_TEMPLATExlsx = '/Dropbox/Phd_R_Ms/PhD_Modeling_DB_GIT/Belmont_AOI_git/Belmont_AOI-runFILES/Network_DF_Period_0t00_TEMPLATE.xlsx'
# PATH_edge_i_cashe_TEMPLATE = '/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/Edge_i_cashe_TEMPLATE.csv' 
PATH_BMAOI_edgeCasheFILES = '/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES'
PATH_Network_DF_Period_0t00_DF = '/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/Network_DF_Period_0t00_DF.csv'

Belmont_Ave=("-12150712#3","-12150712#4","-12150712#6","-12327906#0","-196358954#0","-196358954#3","-196358956#2","-387423966","-423956981","-423956982","-423965484","-423967058#0","-423967058#1","-423967352","-423967353","-423967354","-423967355","-423967356","-423967357","-423967358#1","-423967358#2","-423967359#0","-423967359#1","-424803245","-424803247#1","-424824619","-424824620","-424824621","-424978161","-424978642.0","-424978642.170","-424978643","-424978647#1","-448887867","-448887868","-448887869","-448887870#1","-448887871#0","-448887871#2","-49940170#0","12150712#3","12150712#4","12150712#5","12327906#0","12327906#1","196358954#0","196358954#1","196358956#0","387423966","423956978","423956979","423956980","423965484","423967058#1","423967352","423967353","423967354","423967355","423967356","423967358#0","423967358#2","423967359#0","423967359#1","424803245","424803247#0","424824619","424824620","424824621","424978639.0","424978639.102","424978640","424978643","424978644","424978646","448887867","448887868","448887869","448887870#0","448887871#0","448887871#1")

Belmont_AVEDic ={0: '-12150712#3',1: '-12150712#4',2: '-12150712#6',3: '-12327906#0',4: '-196358954#0',5: '-196358954#3',6: '-196358956#2',7: '-387423966',8: '-423956981',9: '-423956982',10: '-423965484',11: '-423967058#0',12: '-423967058#1',13: '-423967352',14: '-423967353',15: '-423967354',16: '-423967355',17: '-423967356',18: '-423967357',19: '-423967358#1',20: '-423967358#2',21: '-423967359#0',22: '-423967359#1',23: '-424803245',24: '-424803247#1',25: '-424824619',26: '-424824620',27: '-424824621',28: '-424978161',29: '-424978642.0',30: '-424978642.170',31: '-424978643',32: '-424978647#1',33: '-448887867',34: '-448887868',35: '-448887869',36: '-448887870#1',37: '-448887871#0',38: '-448887871#2',39: '-49940170#0',40: '12150712#3',41: '12150712#4',42: '12150712#5',43: '12327906#0',44: '12327906#1',45: '196358954#0',46: '196358954#1',47: '196358956#0',48: '387423966',49: '423956978',50: '423956979',51: '423956980',52: '423965484',53: '423967058#1',54: '423967352',55: '423967353',56: '423967354',57: '423967355',58: '423967356',59: '423967358#0',60: '423967358#2',61: '423967359#0',62: '423967359#1',63: '424803245',64: '424803247#0',65: '424824619',66: '424824620',67: '424824621',68: '424978639.0',69: '424978639.102',70: '424978640',71: '424978643',72: '424978644',73: '424978646',74: '448887867',75: '448887868',76: '448887869',77: '448887870#0',78: '448887871#0',79: '448887871#1'}

equiv_ESAL = {'Car' : 0.0007, 'Truck' : 0.10, 'Semi' : 1.35, '40ftBus' : 1.85}

class Edge():
    # import sumolib
    
    # edge_list = []
    def __init__(self):#, edgeID):#, SUMOID):
        # Edge.edge_list.append(self)
        #set came from here
        self.carCount = 0
        self.truckCount = 0
        self.ESAL_TOT = 0
        self.adaptTraveltime = 0
        self.Effort = 0
        self.Dynamic_Max_Speed = 27.87
        self.meanSpeed = 0
        self.PCR = 0
        self.IRI = 0
        self.AVE_QLTH = 0
        self.ADDT_rand = 0
        self.ADDT_Calc = 0
        self.AGE_0 = 0
        self.SNC = 0
        self.layer_1_dpth_in = 0
        self.layer1_6in_35_50 = 0
        self.layer2_6in_10_25 = 0
        self.layer3_12in_5_17 = 0
        self.ASS_CALI = 0
        self.AGE_t = 0

        #self.Original_Max_Speed = 27.78 # (m/s) or 
        
    def set(self,edgeID, Bel_Dic_ID):
        self.edgeID = edgeID
        self.Bel_Dic_ID = Bel_Dic_ID #Belmont_AVEDic[edgeID]
        self.net = sumolib.net.readNet('C:\Dropbox\Phd_R_Ms\PhD_Modeling_DB_GIT\Belmont_AOI_git\Belmont_AOI-runFILES\Belmount_AOI-V5.net.xml')
        self.vehidLIST = {'vehID_k' : 'veh_k_Type'} # https://stackoverflow.com/questions/1024847/add-new-keys-to-a-dictionary
        self.originalMAXSPEED = 27.87# (m/s) self.net.getEdge(edgeID).getSpeed()
        logger_TEMPDF =pd.read_excel(PATH_Network_DF_Period_0t00_TEMPLATExlsx)
        
        self.AGE_0 = logger_TEMPDF.loc['AGE_0',Bel_Dic_ID]
        self.ADDT_rand = logger_TEMPDF.loc[Bel_Dic_ID,'AADT_rand']
        self.ASS_CALI = logger_TEMPDF.loc[Bel_Dic_ID,'ASS_CALI']
        self.SNC = logger_TEMPDF.loc[Bel_Dic_ID,'SNC']
        self.layer_1_dpth_in = logger_TEMPDF.loc[Bel_Dic_ID,'layer_1_dpth_in']
        self.layer1_6in_35_50 = logger_TEMPDF.loc[Bel_Dic_ID,'layer1_6in_35_50']
        self.layer2_6in_10_25 = logger_TEMPDF.loc[Bel_Dic_ID,'layer2_6in_10_25']
        self.layer3_12in_5_17 = logger_TEMPDF.loc[Bel_Dic_ID,'layer3_12in_5_17']
        self.AADT_Calc = logger_TEMPDF.loc[Bel_Dic_ID,'AADT_rand']
        self.AGE_t = logger_TEMPDF.loc[Bel_Dic_ID,'AGE_0']
    

        
    def make_condition_idex():
#   http://onlinepubs.trb.org/Onlinepubs/trr/1989/1215/1215-001.pdf
#       PCR(t) = 90 - a * [exp(Age^b)-1] * log(ESAL/SNC^c) a = 0.6349; b = 0.4203; and c = 2.7062
        pass
        
        
    def log(self):#, vehID, ESAL_contrib, Emergancy_Stop, Accident):
        self.AGE_t += 1 
        Edge_i_VehIDs_lastStep_j = traci.edge.getLastStepVehicleIDs(self.edgeID)
        # if len(Edge_i_VehIDs_lastStep_j) > 0:
            # print("\n",self.Bel_Dic_ID," - Edge_i_VehIDs_lastStep_j = ",Edge_i_VehIDs_lastStep_j)
        self.AGE_t += 1
        #http://www.sumo.dlr.de/daily/pydoc/traci._edge.html#EdgeDomain-getLastStepMeanSpeed
        for vehID_k in Edge_i_VehIDs_lastStep_j: 
            if vehID_k in self.vehidLIST.keys():
                #print("\nvehID_k is already in here\tvehID_k = ",str(vehID_k))
                continue #?
            else:
                # print("Onward")
                # ##testing vehID_k = Edge_i_VehIDs_lastStep_j[0]
                if re.search('(.*?)Car(.*?)',traci.vehicle.getTypeID(vehID_k),re.IGNORECASE) != None:
                    self.vehidLIST[vehID_k]= traci.vehicle.getTypeID(vehID_k)
                    self.carCount = self.carCount + 1
                    self.ESAL_TOT = self.ESAL_TOT + equiv_ESAL['Car']
                    #print("Car\r\n") #+ edge_i_cashe_csv)
                else:   #Update counts and ESAL Contribution based on type of trcuk ESAL equivalant loads found here http://www.pavementinteractive.org/loads/  
                ##BUT OUTDATED METHOD## New Methods http://www.pavementinteractive.org/equivalent-single-axle-load/
                # https://www.dot.state.pa.us/public/PubsForms/Publications/PUB%20242.pdf - try this out
                    self.vehidLIST[vehID_k]= traci.vehicle.getTypeID(vehID_k)
                    self.truckCount = self.truckCount + 1
                    if re.search('(.*?)Panel(.*)',traci.vehicle.getTypeID(vehID_k),re.IGNORECASE) != None:
                        self.ESAL_TOT = self.ESAL_TOT +  equiv_ESAL['Truck']
                    elif re.search('(.*?)Single_Rear_Truck(.*)',traci.vehicle.getTypeID(vehID_k),re.IGNORECASE) != None:
                        self.ESAL_TOT = self.ESAL_TOT +  equiv_ESAL['Truck']
                    elif re.search('(.*?)Double_Rear_Truck(.*)',traci.vehicle.getTypeID(vehID_k),re.IGNORECASE) != None:
                        self.ESAL_TOT = self.ESAL_TOT +  equiv_ESAL['Semi']
                    elif re.search('(.*?)bus(.*)',traci.vehicle.getTypeID(vehID_k),re.IGNORECASE) != None:
                        self.ESAL_TOT = self.ESAL_TOT +  equiv_ESAL['40ftBus']
                    else:
                        print("I Don't Know What you Are")

                        
    def getTravel_times_n_Effort(self):
        self.adaptTraveltime = traci.edge.getAdaptedTraveltime(self.edgeID,int(traci.simulation.getCurrentTime()/1000))
        self.Effort = traci.edge.getEffort(self.edgeID,int(traci.simulation.getCurrentTime()/1000)) #5.1 said you broke it
        self.meanSpeed = traci.edge.getLastStepMeanSpeed(self.edgeID)
        #http://www.sumo.dlr.de/daily/pydoc/traci._edge.html#EdgeDomain-getLastStepMeanSpeed
        # if int(traci.edge.getAdaptedTraveltime(self.edgeID,20)) == int(-1.0):
            # self.adaptTraveltime = traci.edge.adaptTraveltime(self.edgeID,1)
        # else:
            # self.adaptTraveltime = traci.edge.getAdaptedTraveltime(self.edgeID,20)
        # # print(traci.edge.getAdaptedTraveltime(self.edgeID,int(traci.simulation.getCurrentTime()/1000)))
        # if int(traci.edge.getEffort(self.edgeID,20)) == int(-1.0):
            # self.Effort = traci.edge.setEffort(self.edgeID,1)
        # else:
            # self.Effort = traci.edge.getEffort(self.edgeID,int(traci.simulation.getCurrentTime()/1000))
        # print(traci.edge.getEffort(self.edgeID,traci.simulation.getCurrentTime()/1000))
        
    @classmethod # means class instead of self (Edge.create_Edge_Instances() becomes create_Edge_Instances(Edge))
    def create_Edge_Instances(cls):
        # @staticmethod # means no self
        # def create_Edge_Instances():
        tabber = "<>"
        print("\n\t\t\t\t<><>Please wait creating Edge Instances<><>\n")
        edgeLISTa = list()
        counter = 0
        for edge_i in Belmont_Ave[:]: #### HERE TO INCLUDE ALL INSTANCES
            edgeNAME = "edge_"+str(counter) #str(SP.Belmont_AVEDic[counter])
            edgeLISTa.append(edgeNAME)
            # print(edgeLISTa[counter])
            edgeLISTa[counter]= cls()
            edgeLISTa[counter].set(Belmont_Ave[counter],counter)
            edgeLISTa[counter].getTravel_times_n_Effort()
            print("Edge = ", counter,tabber,end='\r',flush=True)#, edgeLISTa[counter].__dict__)
            counter = counter + 1
            # tabber = str(tabber+"<>")#\t")
            if (counter/47).is_integer() == True:
                tabber = "\n<>"
            else:
                tabber = str(tabber+"<>")#\t")
        print("\n\t\t\t\t<<<Edge Instances loaded on edgeLISTa>>>\n")
        print("Testing edgeLISTa[4].__dict__ ...\n",edgeLISTa[4].__dict__,"\n")
        return edgeLISTa
        
    # def __str__(self): #[str(edge) for edge in edgeLISTa]
        # return str(self.__dict__)
    
    # def __repr__(self):
        # return repr(self.__dict__)
        
    def _resetLOGGER(self,edgeLISTa):
        print("\n\nReset_Edge_CasheFiles has started")
        counter = 0
        for edge_i in Belmont_Ave[:]:
        # edgeLISTa[Belmont_Ave.index(self.edgeID)].vehidLIST= {'vehID_k' : 'veh_k_Type'}
            edgeLISTa[counter].vehidLIST= {'vehID_k' : 'veh_k_Type'}
            counter = counter + 1
        return #print("Reset_Edge_CasheFiles has completed\n\n")

    def setNewMaxSpeed(edgeLISTa, logger_TEMPDF):
        print("\n\n<><><><Changing SUMO EGDE MAX SPEED><><><>\n")
        counter = 0
        for rd in Belmont_Ave:
            # print("Please be true ", int(logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(rd),'Dynamic_Max_Speed']) == edgeLISTa[rd].Dynamic_Max_Speed)
            maxSpeed_i = logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(rd),'Dynamic_Max_Speed']
            maxSpeed_i = round(maxSpeed_i.iloc[0],2)
            # maxSpeed_i = edgeLISTa[rd].Dynamic_Max_Speed
            traci.edge.setMaxSpeed(rd,maxSpeed_i)
            edgeLISTa[counter].Dynamic_Max_Speed = maxSpeed_i
            counter += 1
            if maxSpeed_i <= 27:
                print("The new speed for edge ",rd," is now ",round(maxSpeed_i.iloc[0],2))
        ### Need a new method to change driver imperfection based on road damage
        ## traci.vehicle.setImperfection(vehID..., function relating road index to driver imperfection)
        
        
    def getMaxSpeed(edgeLISTa):#, logger_TEMPDF):
        # current_time = str(traci.simulation.getCurrentTime()/1000)
        # print("\n[[[[MAX SPEEDS]]]]\t\tCurrent Time:",current_time,"\n")
        # edge_i_speed_t = 0
        # counter = 0
        # for edge_i in Belmont_Ave[:]:
            # edgeLISTa[counter].Dynamic_Max_Speed = edge_i_speed_t
            # print(" ", edgeLISTa[counter].edgeID," : ", edge_i_speed_t,"; ", end='\r', flush=True)
            # counter = counter + 1
        return print("<>\n<>><<><><><><\n\n\n\<><><><\t\t<><><><><><\t\tI'm a failure... getMaxSpeed<>\n<>><<><><><><\n\n\n\<><><><\t\t<><><><><><")
       
class Network_Period:

    def load_n_create_Excel_NetworkFile(SUMO_outPUT_PREFIX, PERIOD_VARRIABLE,steps_TT, display = None, PATH=None):
        if display != 0:
            print("Loading and creating Excel Network File", "SUMO_outPUT_PREFIX = ", SUMO_outPUT_PREFIX)
        if PATH == None:
            PATH_Network_DF_Period_0t00_TEMPLATExlsx = '/Dropbox/Phd_R_Ms/PhD_Modeling_DB_GIT/Belmont_AOI_git/Belmont_AOI-runFILES/Network_DF_Period_0t00_TEMPLATE.xlsx'
            PATH_to_Save_to = "/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES/Network_" +SUMO_outPUT_PREFIX + "_PeriodBook.xlsx"
        wb = OPENxlsx.Workbook()
        full_Run_Time = 90000 #steps_TT
        Network_DF_Period_0t00xlsx=pd.read_excel(PATH_Network_DF_Period_0t00_TEMPLATExlsx)
        periodNamesLISTa = list()
        for period in range(int(int(full_Run_Time)/PERIOD_VARRIABLE)):
            periodNAME = "Period_"+str(((PERIOD_VARRIABLE)*(period))+1)+"_to_"+str((((PERIOD_VARRIABLE)*(period)+1)+PERIOD_VARRIABLE))
            periodNamesLISTa.append(periodNAME)
        # print("/n<><><periodNamesLISTa = ",periodNamesLISTa)
        for i in range(0,len(periodNamesLISTa[:])):
            # print("\nwb.get_sheet_names()[0] = ",wb.get_sheet_names()[0])
            ws = wb.create_sheet(title = periodNamesLISTa[i])
        if wb.get_sheet_names()[0] == 'Sheet':
            print("\nwb.get_sheet_names()[0] = ",wb.get_sheet_names()[0])
            delme = wb.get_sheet_by_name('Sheet')
            print("\nRemoving wb.get_sheet_names()[0] = ",wb.get_sheet_names()[0])
            wb.remove_sheet(delme)
            wb.save(filename = PATH_to_Save_to)
        wb.save(filename = PATH_to_Save_to)
        wb.close()
        return wb, periodNamesLISTa 
        
    def fillOutworksheet(SUMO_outPUT_PREFIX,periodCounter,edgeLISTa):#,periodNamesLISTa):
        #https://chrisalbon.com/python/data_wrangling/pandas_dataframe_load_xls/
        #### BIG CONTRIBUTION PART
        if periodCounter == 0:
            return
        # PATH_Network_DF_Period_0t00_TEMPLATExlsx = '/Dropbox/Phd_R_Ms/PhD_Modeling_DB_GIT/Belmont_AOI_git/Belmont_AOI-runFILES/Network_DF_Period_0t00_TEMPLATE.xlsx'
        print("\n\n<><><periodCounter = ",int(round(periodCounter)))
        periodCounter = int(round(periodCounter))#periodCounter = int(round(periodCounter)-1) # "-1" because we want to fill in the sheet from the perivious period.
        logger_TEMPDF =pd.read_excel(PATH_Network_DF_Period_0t00_TEMPLATExlsx) #creates a template
        # logger_TEMPDF =pd.read_excel(SP.PATH_Network_DF_Period_0t00_TEMPLATExlsx)
        counter = 0
        #for edge_i in Belmont_Ave[:]:
        for n in range(len(edgeLISTa)): ## I need Belmont_AVEDic[n] == edgeLISTa[n].edgeID == logger_TEMPDF.loc[n,'Belmont_AVEDic_ID']. 1-8-18 saved over tempalte with for loop range(len(edgeLISTa)): newTemplate.loc[i,'Belmont_AVEDic_ID'] = edgeLISTa[i].edgeID 
            logger_TEMPDF.loc[n,'Total_Vehicles'] = (edgeLISTa[n].truckCount + edgeLISTa[n].carCount)
            logger_TEMPDF.loc[n,'Total_Trucks'] = edgeLISTa[n].truckCount
            ## logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(Belmont_AVEDic[n]),'Total_Trucks'] = edgeLISTa[n].truckCount
            logger_TEMPDF.loc[n,'ESAL_TOT'] = edgeLISTa[n].ESAL_TOT
            Condition_RTi = 100.00 - (edgeLISTa[n].ESAL_TOT) * 0.01339
            logger_TEMPDF.loc[n,'Condition_Index'] = Condition_RTi
            # maxSpeedo = logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(Belmont_AVEDic[n]),'Original_Max_Speed']
            maxSpeedo = edgeLISTa[n].originalMAXSPEED
            maxSpeed_i =(maxSpeedo - (maxSpeedo**((100-Condition_RTi)/96))+4.5675) #Units m/s
            logger_TEMPDF.loc[n,'Dynamic_Max_Speed'] = maxSpeed_i
            edgeLISTa[n].Dynamic_Max_Speed = maxSpeed_i
        ## Needed for metrics
            edgeLISTa[n].meanSpeed = int(traci.edge.getLastStepMeanSpeed(edgeLISTa[n].edgeID))
            edgeLISTa[n].PCR = 90-0.6349 * (np.exp((edgeLISTa[n].AGE_0+edgeLISTa[n].AGE_t/31556926)**0.4203)-1) * np.log(edgeLISTa[n].ESAL_TOT/((edgeLISTa[n].SNC)**2.7062))
            edgeLISTa[n].IRI = 52 + 8.1 * ((int(edgeLISTa[n].AGE_t)/31556926))+0.0009*edgeLISTa[n].ADDT_Calc
            edgeLISTa[n].AVE_QLTH = 0
            edgeLISTa[n].OCPNY = traci.edge.getLastStepOccupancy(edgeLISTa[n].edgeID)
            # ##Validating writing below here
            # if n == 4:
                # print("\n[<<[<[<>]>]>>]\nedgeLISTa[4].truckCount = ",edgeLISTa[4].truckCount,"<><>\nlogger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edgeLISTa[n].edgeID),'Total_Trucks']= ",logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edgeLISTa[n].edgeID),'Total_Trucks'],"\n[<<[<[<>]>]>>]\n")
            ## Do I need to reset the logger? Will it slow things down towards the end? 
            if len(edgeLISTa[n].vehidLIST) > 100:
                if n == 48:
                    memoryUse = py.memory_info()[0]/2.**30  # memory use in GB...I think
                    print('memory use:', memoryUse)
                    edgeLISTa[n]._resetLOGGER(edgeLISTa)
                    memoryUse = py.memory_info()[0]/2.**30  # memory use in GB...I think
                    print('memory use:', memoryUse)
                edgeLISTa[n]._resetLOGGER(edgeLISTa)

        ### NOW SAVE THIS DATAFRAME [[logger_TEMPDF]] TO THE SHEET THAT IT CAME FROM IN THE WORKBOOKER
        Network_Period.myWrite_to_excel(logger_TEMPDF,periodCounter,SUMO_outPUT_PREFIX,edgeLISTa)
        ### Change max speed of edge based on roadway damage
        Edge.setNewMaxSpeed(edgeLISTa,logger_TEMPDF)
        Edge.getMaxSpeed(edgeLISTa)
        print("Testing edgeLISTa[48].__dict__ ...\n",edgeLISTa[48].__dict__,"\n======",SUMO_outPUT_PREFIX,"======")
        print("\t\t\t\t======",SUMO_outPUT_PREFIX,"======")


        return  
            
    def myWrite_to_excel(logger_TEMPDF,periodCounter,SUMO_outPUT_PREFIX,edgeLISTa):
        # https://openpyxl.readthedocs.io/en/default/tutorial.html#data-storage   
        # http://openpyxl.readthedocs.io/en/default/pandas.html#working-with-pandas-dataframes
        ### TRY THIS http://pbpython.com/improve-pandas-excel-output.html and then this https://codereview.stackexchange.com/questions/180405/writing-excelsheet-from-python-dataframe
        # from openpyxl import load_workbook
        ## openpyxl.worksheet.Worksheet.cell() method from  
        PATH_to_Save_to = "/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES/Network_" +SUMO_outPUT_PREFIX + "_PeriodBook.xlsx"
        wb =  OPENxlsx.load_workbook(filename = PATH_to_Save_to)
        periodCounter = int(round(periodCounter))
        ws = wb[wb.get_sheet_names()[periodCounter-1]]
        print("\n[[[<><><>]]]\nStarting myWrite_to_excel()\n\t PATH_to_Save_to: ",PATH_to_Save_to,"\n\t Current Worksheet: ",ws,"\n\n[[[<><><>]]]")
        wsROW = 0
        for i in range(logger_TEMPDF.shape[0]):
            wsROW +=1 
            wsCOL =0
            for j in range(logger_TEMPDF.shape[1]):
                wsCOL +=1
                if i == 1:
                    ws.cell(row=wsROW, column=wsCOL, value=logger_TEMPDF.columns[j])
                    if i == 1 and j == 1:
                        print("\t\t\t\t<Writing Headers>")
                    ws.cell(row=wsROW+1, column=wsCOL, value=logger_TEMPDF.iloc[i,j])
                else:
                    ws.cell(row=wsROW+1, column=wsCOL, value=logger_TEMPDF.iloc[i,j])
            if i == 7:
                wsdf = pd.DataFrame(ws.values)
                print("\nPlease be true ... ", wsdf.iloc[7,6] == edgeLISTa[6].truckCount)
                logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edgeLISTa[6].edgeID),'Total_Trucks']
        wsdf = pd.DataFrame(ws.values)
        wb.save(PATH_to_Save_to)
        print("\n>>>Data written to:: wb[wb.get_sheet_names()[periodCounter-1]]= ",wb[wb.get_sheet_names()[periodCounter-1]],"\n\t\t\t\t\t this should match ws: ",ws,"\n[[[[<<<<Next Period>>>>]]]]: ",wb.get_sheet_names()[periodCounter])
        return wb
        
       
class Initializer:
    
    def __init__(self,new):
        return #new
        
        
    def inputPeriod_asNumber(new):
        #global PERIOD_VARRIABLE
        if new ==1:
            while True: #http://www.101computing.net/number-only/
                    try:
                        userINPUT = input("\r\n\nPress [1] to enter a user defined Period_Interval.\r\nPress anything else to continue default is seconds 3600 or 1 hour? ... ") 
                        if  userINPUT != '1':
                            PERIOD_VARRIABLE = 3600 #Default to 1 hour
                            print("\n\nPERIOD_VARRIABLE = ", PERIOD_VARRIABLE,"\n")
                            return int(PERIOD_VARRIABLE)
                            break
                        else:
                            PERIOD_VARRIABLE = int(input("\nWhat would you like the Period of Collection to be default is seconds 3600 or 1 hour? ... "))
                    except ValueError:# as err: #https://docs.python.org/3/library/exceptions.html#BaseException 
                        print("\n!!!!!!!!!!!!Type ERROR\n\t\tPlease Enter an Integer\n\tWhat would you like the Period of Collection to be default is seconds 3600 or 1 hour? ... ")
                        continue
                    else:
                        print("PERIOD_VARRIABLE = ", PERIOD_VARRIABLE)
                        return int(PERIOD_VARRIABLE)
                        break
        else:
            print("PERIOD_VARRIABLE = ", PERIOD_VARRIABLE)
        return int(PERIOD_VARRIABLE)


    def startSUMO(SUMO_Traci_PORT,useCase=None):
        sumoCmd #= [sumoBinary, "-c", configPATH, "--start"]
        sumoGUICmd #= [sumoGUIBinary, "-c", configPATH, "--start"]
        portTouse = str(SUMO_Traci_PORT)
        # global Start_Time
        # while rideAgain == 1:
            #print("useCase =",useCase)
        if useCase == None:
           print("\t\t\t\t\t\t###STARTING####")
           useCase = input("\n\tPress 1 for a new Run:  \n") #\nPress Anything to continue...
        if useCase == "1":
            GUI_01 = input("Press Zero to use GUI? (0)\n")
            inputPort = input("Confirm with: BMAOI-TRACI.sumocfg .. port in File: .."+ str(SUMO_Traci_PORT)+"\nPlease select a port...\t") # 
            if inputPort == '':
                portTouse = int(SUMO_Traci_PORT)
            else:
                portTouse = inputPort
            print("\n\t\tTrying port ..  ",portTouse,"\n\n")
            if str(GUI_01) != "0":
                traci.start(sumoCmd,int(portTouse))
            else:
                traci.start(sumoGUICmd,int(portTouse))
            traci.simulationStep()
            Start_Time = int(traci.simulation.getCurrentTime()/1000)
            #useCase == "Continue"
        elif useCase == "Continue":
            print("\t\t\t\t   !!!!!!!!!!!!!!CONTINUING!!!!!!!!!!!!!\t\t\t")
            typeRun = "4"
            Start_Time = int(traci.simulation.getCurrentTime()/1000)
        elif useCase == "x":
            PC.condition.DisplayFiles()
            traci.close()
        print("\nuseCase = ",useCase)
        return Start_Time, portTouse, useCase

        
class Runner:
    # def __init__(PERIOD_VARRIABLE,periodCounter=None):
        # self.PERIOD_VARRIABLE = PERIOD_VARRIABLE
        # self.periodDIC = {}
        # if periodCounter == None:
            # self.periodDIC={'period_1.0' : Network_Period(self.PERIOD_VARRIABLE,SUMO_outPUT_PREFIX)}
        # else:
            # self.periodDIC={"period_"+str(periodCounter):Network_Period(PERIOD_VARRIABLE,SUMO_outPUT_PREFIX)}

    def runtypeAsker(typeRun=None):
        print("\nBeginning Journey Now\n\tPlease See Gui for Show\n")
        typeRun = input("\n\nPush T to Run Simulation until the end\nPush 2 to run for XXX Steps\nPush 3 to run for X number of periods\nPush 4 to run until input time\nCurrent Step_Time : "+str(traci.simulation.getCurrentTime()/1000)+"\n\n\tEnter Run Type Here: ")
        if typeRun == '':
            typeRun = "2"
            print("You didn't specify! Going with default value... typeRun =", str(typeRun))
        else:
            print("You pressed", str(typeRun),"= typeRun")
        return typeRun
    
        ###Run the Code####
    def releaseTraci(Start_Time,typeRun,edgeLISTa,PERIOD_VARRIABLE,SUMO_outPUT_PREFIX,periodNamesLISTa,steps_TT=None):
        print("\t\t\t\t======",SUMO_outPUT_PREFIX,"======")
        #Defining the Constants
        if steps_TT == None:
            steps_TT = 0
        stepCounter = 0
        addTesterTrucks = 0
        #Initializer.runSUMO(SUMO_Traci_PORT)[0]
        # Period_Break_Point=('50','1000','2500')
        periodCounter = (traci.simulation.getCurrentTime()/1000)#/PERIOD_VARRIABLE # was 0
        # if typeRun == None:
            # typeRun = 2
        if str(typeRun) == "T": #Run Simulation unti the end
            while traci.simulation.getMinExpectedNumber() > 0:
                #print("step No. = ", step," out of ",No_next_steps) #Showing the steps boggy downy machiny
                traci.simulationStep()
                Runner.thingsTodoWhileStepping(edgeLISTa,PERIOD_VARRIABLE,periodCounter,SUMO_outPUT_PREFIX,periodNamesLISTa)
                periodCounter += 1
        elif str(typeRun) == "2": 
            steps_TT = input("How many steps would you like to take? ")
            #No_next_steps = int(steps_TT)
            if addTesterTrucks == "1":
                traci.simulationStep()
                general.addBen_Trackers()
            for step in range(int(steps_TT)):
                traci.simulationStep()
                Runner.thingsTodoWhileStepping(edgeLISTa,PERIOD_VARRIABLE,periodCounter,SUMO_outPUT_PREFIX,periodNamesLISTa)
                periodCounter += 1#(1/PERIOD_VARRIABLE)
        elif str(typeRun) == "3": # run for X number of periods
            numb_Periods = input("How many periods would you like to run for? ")
            steps_TT = int(PERIOD_VARRIABLE * int(numb_Periods))
            next_sim_stop_time = int((traci.simulation.getCurrentTime()/1000)+steps_TT)
            print("Simulation set to run until ",str(next_sim_stop_time),"\n") 
            for step in range(int(steps_TT+1)):
                traci.simulationStep()
                Runner.thingsTodoWhileStepping(edgeLISTa,PERIOD_VARRIABLE,periodCounter,SUMO_outPUT_PREFIX,periodNamesLISTa)
                periodCounter += 1#(1/PERIOD_VARRIABLE)
        elif str(typeRun) == "4": #run until input time VAR -> NextTimeToStop			
            print("\nCurrent Time is: ",str(traci.simulation.getCurrentTime()/1000))
            # import ipdb
            # ipdb.set_trace()
            NextTimeToStop = input("\nWhat time (in seconds) would you like this run of the simulation to end at? ... ")
            NextTimeToStop = int(NextTimeToStop)
            print("\t\t\t<><>Running until time: ",NextTimeToStop,"<><>")
            while traci.simulation.getCurrentTime()/1000 < NextTimeToStop:
                traci.simulationStep()
                Runner.thingsTodoWhileStepping(edgeLISTa,PERIOD_VARRIABLE,periodCounter,SUMO_outPUT_PREFIX,periodNamesLISTa)
                periodCounter += 1       #(1/PERIOD_VARRIABLE)
        print("<>\n<><>\n<><><>\nperiodCounter = ",periodCounter)#, "periodNamesLISTa = ",periodNamesLISTa)
        return # edgeLISTa
                    
    def thingsTodoWhileStepping(edgeLISTa,PERIOD_VARRIABLE,periodCounter,SUMO_outPUT_PREFIX,periodNamesLISTa):
        for i in range(len(Belmont_Ave)):
            edgeLISTa[i].log()
        if (periodCounter/PERIOD_VARRIABLE).is_integer():
            periodCounter = (periodCounter/PERIOD_VARRIABLE)
            print("\t\t\t\t======",SUMO_outPUT_PREFIX,"======")
            Network_Period.fillOutworksheet(SUMO_outPUT_PREFIX,periodCounter,edgeLISTa)
                    

    def other_parts_not_yet_implemented():
        ## General Output and things after each run has done its things
        #Show Network files
        #print("Network_DF_Period_PATH = ",Network_DF_Period_PATH, "Network_DF_Period_DF = ", Network_DF_Period_DF)
        PC.condition.DisplayFiles(PERIOD_VARRIABLE)
        # print("\n\nThanks for riding the SUMO Traci Train. Please come back again soon\r\n$ run /GitHub/PhD_Modeling/PYTHON_GIT_SUMO/SUMO_Runner.py\n")
        # goAgain = input("\n\n >>>>>>> This part of the journey has ended. Press anything to contine?\n>>>>>>>>>Or x to EXIT\n\n>  >\t> >\t>> ")
        # if goAgain == "x":
# #            traci.close()
            # print("\n\n^^---Journey has ended---^^\n\n") #,edge_VehIDhistoryPD,"
        # else:
            # #general.addNewColumn(2,Period_Time)
            # general.releaseTraci(useCase="Continue")
        general.outPUT(useCase,Start_Time,PERIOD_VARRIABLE, steps_TT)
        
class RunFileInfo:

    def __init__(self):#, SUMO_outPUT_PREFIX,SUMO_Traci_PORT):
        pass
        
    def GetSimulationRunPrefix(display):
        import re
        global SUMO_outPUT_PREFIX
        global SUMO_Traci_PORT
        with open('/Dropbox/Phd_R_Ms/PhD_Modeling_DB_GIT/Belmont_AOI_git/Belmont_AOI-runFILES/BMAOI-TRACI.sumocfg') as f:
            termLIST = list()
            termcounter = 0
            for line in f:
                # print("SUMO_outPUT_PREFIX = ",SUMO_outPUT_PREFIX)
                # if SUMO_outPUT_PREFIX =="":
                if '<output-prefix value="' in line:
                    print(line)
                    SUMO_outPUT_PREFIX_LINE = line
                    SUMO_outPUT_PREFIX = re.search('<output-prefix value="(.*)-(.*)-"/>',SUMO_outPUT_PREFIX_LINE, re.IGNORECASE).group(2)
                    termLIST.append(re.search('<output-prefix value="(.*)-(.*)-"/>',SUMO_outPUT_PREFIX_LINE, re.IGNORECASE).group(2))
                    SUMO_outPUT_PREFIX = termLIST[0]
                    # print("\nSUMO_outPUT_PREFIX = ",SUMO_outPUT_PREFIX,"\n")
                    termcounter +=1
                    break
            for line in f:
                if '<remote-port value=' in line:
                    SUMO_outPUT_Port_LINE = line
                    SUMO_Traci_PORT = re.search('<remote-port value="(.*)"/>',SUMO_outPUT_Port_LINE, re.IGNORECASE).group(1)
                    termLIST.append(re.search('<remote-port value="(.*)"/>',SUMO_outPUT_Port_LINE, re.IGNORECASE).group(1))
        if display == 1:
            print("\nSUMO_outPUT_PREFIX = ", SUMO_outPUT_PREFIX,"\n\SUMO_Traci_PORT = ", SUMO_Traci_PORT)

        return SUMO_outPUT_PREFIX, SUMO_Traci_PORT#, print("\nSUMO_outPUT_PREFIX = ", SUMO_outPUT_PREFIX,"\nSUMO_Traci_PORT = ", SUMO_Traci_PORT)
        
print("Your new sumoPython_git_A has been loaded!")
        ## Python Sumo Script

## Grave yard

