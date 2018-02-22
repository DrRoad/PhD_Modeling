class Network_Period:
    def load_n_create_Excel_NetworkFile(SUMO_outPUT_PREFIX, PERIOD_VARRIABLE,steps_TT, display = None, PATH=None):
        if display != 0:
            print("Loading and creating Excel Network File", "SUMO_outPUT_PREFIX = ", SUMO_outPUT_PREFIX)
        if PATH == None:
            PATH_Network_DF_Period_0t00_TEMPLATExlsx = '/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES/Network_DF_Period_0t00_TEMPLATE.xlsx'
            PATH_to_Save_to = "/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES/Network_" +SUMO_outPUT_PREFIX + "_PeriodBook.xlsx"
        wb = OPENxlsx.Workbook()
        # worksheet1 = testXLSX.active
        # worksheet1.title = "work sheet number 1"
        # worksheet_2 = testXLSX.create_sheet(title = "WS#2")

        estimated_Run_Time = steps_TT
        Network_DF_Period_0t00xlsx=pd.read_excel(PATH_Network_DF_Period_0t00_TEMPLATExlsx)
        periodNamesLISTa = list()
        for period in range(int(int(estimated_Run_Time)/PERIOD_VARRIABLE)):
            periodNAME = "Period_"+str(((PERIOD_VARRIABLE)*(period))+1)+"_to_"+str((((PERIOD_VARRIABLE)*(period)+1)+PERIOD_VARRIABLE))
            periodNamesLISTa.append(periodNAME)
        print("/n<><><periodNamesLISTa = ",periodNamesLISTa)
        from openpyxl.utils.dataframe import dataframe_to_rows
        for i in range(0,len(periodNamesLISTa[:])):
            # print("\nwb.get_sheet_names()[0] = ",wb.get_sheet_names()[0])
            ws = wb.create_sheet(title = periodNamesLISTa[i])
        if wb.get_sheet_names()[0] == 'Sheet':
            print("\nwb.get_sheet_names()[0] = ",wb.get_sheet_names()[0])
            delme = wb.get_sheet_by_name('Sheet')
            print("\nRemoving wb.get_sheet_names()[0] = ",wb.get_sheet_names()[0])
            wb.remove_sheet(delme)
            wb.save(filename = PATH_to_Save_to)
            # for r in dataframe_to_rows(Network_DF_Period_0t00xlsx, index=False, header=True):
                # wb[wb.get_sheet_names()[i]].append(r)
        
        wb.save(filename = PATH_to_Save_to)
        return wb, periodNamesLISTa 
        
    def fillOutworksheet(SUMO_outPUT_PREFIX,periodCounter,edgeLISTa,periodNamesLISTa):
        #https://chrisalbon.com/python/pandas_dataframe_load_xls.html
        if periodCounter == 0:
            return
        # from openpyxl import load_workbook
        PATH_Network_DF_Period_0t00_TEMPLATExlsx = '/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES/Network_DF_Period_0t00_TEMPLATE.xlsx'
        PATH_to_Save_to = "/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES/Network_" +SUMO_outPUT_PREFIX + "_PeriodBook.xlsx"
        wb =  OPENxlsx.load_workbook(filename = PATH_to_Save_to)
        periodCounter = int(round(periodCounter))
        print("\n\n<><><periodCounter = ",int(round(periodCounter)))#,">>>\wb[wb.get_sheet_names()[periodCounter+1]] = ",wb[wb.get_sheet_names()[periodCounter+1]])#, " ...But...\nNow periodCounter = ", round(periodCounter))
        periodCounter = int(round(periodCounter))#periodCounter = int(round(periodCounter)-1) # "-1" because we want to fill in the sheet from the perivious period.
        logger_TEMPDF =pd.read_excel(PATH_Network_DF_Period_0t00_TEMPLATExlsx) #grabs the currennt period sheet
        counter = 0
        for edge_i in Belmont_Ave[:]: 
            logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edge_i),'Total_Vehicles'] = (edgeLISTa[counter].truckCount + edgeLISTa[counter].carCount)
            logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edge_i),'Total_Trucks'] = edgeLISTa[counter].truckCount
            logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edge_i),'ESAL'] = edgeLISTa[counter].ESAL_TOT
            Condition_RTi = 100.00 - (edgeLISTa[counter].ESAL_TOT) * 0.01339
            logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edge_i),'Condition_Index'] = Condition_RTi
            # maxSpeedo = logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edge_i),'Original_Max_Speed']
            maxSpeedo = edgeLISTa[counter].originalMAXSPEED
            maxSpeed_i =(maxSpeedo - (maxSpeedo**((100-Condition_RTi)/96))+4.5675) #Units m/s
            logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edge_i),'Dynamic_Max_Speed'] = maxSpeed_i
            counter += 1
        ### NOW I WANT TO SAVE THIS DATAFRAME [[CURRENT_SHEET]] TO THE SHEET THAT IT CAME FROM IN THE WORKBOOKER
        # print("\n logger_TEMPDF = \n",logger_TEMPDF, "\n<><>\n")
        from openpyxl.utils.dataframe import dataframe_to_rows  ##http://openpyxl.readthedocs.io/en/default/tutorial.html
        # # # # # # # # # # # # Create a Pandas Excel writer using XlsxWriter as the engine.
        # # # # # # # # # # # # writer = pd.ExcelWriter(PATH_to_Save_to, engine='openpyxl')
        # # # # # # # # # # # # logger_TEMPDF.to_excel(writer, sheet_name =periodNamesLISTa[periodCounter])
        # http://openpyxl.readthedocs.io/en/default/pandas.html#working-with-pandas-dataframes
        for r in dataframe_to_rows(logger_TEMPDF, index=False, header=True): 
            wb[wb.get_sheet_names()[periodCounter-1]].append(r) # "-1" because we want to fill in the sheet from the perivious period.
        print("\n>>>Data written to:: wb[wb.get_sheet_names()[periodCounter]] = ",wb[wb.get_sheet_names()[periodCounter-1]],"\n")
        # logger_TEMPDF.to_excel(writer, sheet_name =periodNamesLISTa[periodCounter])# 
        # print(">>>\nwb.get_sheet_names() = ",wb.get_sheet_names())
        wb.save(filename = PATH_to_Save_to)
        Edge.setNewMaxSpeed(edgeLISTa,logger_TEMPDF)
        ## Why do I need to reset the logger? Will it slow things down towards the end? 
        ## Edge._restLOGGER(edgeLISTa)
        return wb




#https://stackoverflow.com/questions/42370977/how-to-save-a-new-sheet-in-an-existing-excel-file-using-pandas
import pandas as pd
import numpy as np
from openpyxl import load_workbook

PATH_Network_DF_Period_0t00_TEMPLATExlsx = '/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES/Network_DF_Period_0t00_TEMPLATE.xlsx'
SUMO_outPUT_PREFIX = "Run_xx"

PATH_to_Save_to = "/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES/Network_" +SUMO_outPUT_PREFIX + "_PeriodBook.xlsx"

workBooker = load_workbook(PATH_Network_DF_Period_0t00_TEMPLATExlsx)
writer = pd.ExcelWriter(PATH_to_Save_to, engine = 'openpyxl')
writer.workBooker = workBooker
Network_DF_Period_0t00xlsx=pd.read_excel(PATH_Network_DF_Period_0t00_TEMPLATExlsx)
currentNetwork_DF_FILE = Network_DF_Period_0t00xlsx
for i in range(10):
    currentNetwork_DF_FILE.to_excel(writer, sheet_name = ('period_'+str(i)))
writer.save()
writer.close()


