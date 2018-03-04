## Last worked on 2/23/2018


## GIT COMMIT AND PUSH
## GIT COMMIT AND PUSH # python run /Dropbox/Phd_R_Ms/PhD_Modeling_DB_GIT/PYTHON_DB_GIT_SUMO/Python_Sumo_Script.py ## GIT COMMIT AND PUSH
## GIT COMMIT AND PUSH
## I think best just to run with Jupyter Notebook... for debugging.

##Note to check: https://intelligea.wordpress.com/2015/08/05/check-if-python-version-is-64-or-32-bit/
# import struct;print struct.calcsize("P") * 8 - yes 64bit


## GIT COMMIT AND PUSH


# Ask for Parameters
    # Period Length...
# import os, sys
import numpy as np
import traci
import traci.constants as tc
import re
import time
import sumolib
# import cProfile, pstats , io
import pandas as pd
import sumoPython_git_A as SP
import openpyxl as OPENxlsx
# pd.set_option('display.max_rows', 5)

# loading_FROM = input("\n\t\t<><><> press d to run from Drexel location <><><>")
# print("loading_FROM = ",loading_FROM)
# if loading_FROM == str("d"):
    # configPATH = "C:/Dropbox/Phd_R_Ms/PhD_Modeling_DB_GIT/Belmont_AOI_git/Belmont_AOI-runFILES/BMAOI_TRACI_DXL.sumocfg"
# else:
# Run_list_TXT = '/Dropbox/Phd_R_Ms/PhD_Modeling_DB_GIT/Belmont_AOI_git/Belmont_AOI-runFILES/Run_list_FILE.txt'
# with open(Run_list_TXT,'a') as f:
    # XMLdataRL = open(Run_list_TXT).readlines()
    # run_LIST = list()
    # for line in range(len(XMLdataRL)):
        # if '00' in XMLdataRL[line]:
            # run_LIST.append(int(XMLdataRL[line]))
        # else:
            # next
    # print(max(run_LIST))
    # new_run = '\n00'+str(max(run_LIST)+1)
    # f.write(new_run)
# print(new_run)
configPATH_MASTER = "C:/Dropbox/Phd_R_Ms/PhD_Modeling_DB_GIT/Belmont_AOI_git/Belmont_AOI-runFILES/BMAOI_TRACI_DB.sumocfg"
configPATH = "C:/Dropbox/Phd_R_Ms/PhD_Modeling_DB_GIT/Belmont_AOI_git/Belmont_AOI-runFILES/BMAOI_TRACI_DB.sumocfg"

# %colors Linux
PERIOD_VARRIABLE = 3600#SP.Initializer.inputPeriod_asNumber(new=1)
fileINFO = SP.RunFileInfo.GetSimulationRunPrefix(display=1,prefix=1,port=1)
SUMO_outPUT_PREFIX = SP.RunFileInfo.GetSimulationRunPrefix(display=0,prefix=1,port=0)
SUMO_Traci_PORT = int(SP.RunFileInfo.GetSimulationRunPrefix(display=0,prefix=0,port=1))
SP.Initializer.startSUMO(SUMO_Traci_PORT,useCase=str(1),GUI_01=0)
    # Ask for Steps to take or Time to run until
typeRun = '2'#SP.Runner.runtypeAsker()
Start_Time = int(traci.simulation.getCurrentTime()/1000) 
print("======",SUMO_outPUT_PREFIX,"======")
#SP.Initializer.runSUMO(SUMO_Traci_PORT,useCase="Continue")[0]
# estimated_Run_Time = input("...\n\n\nHow long will you run this file for... ")
# if estimated_Run_Time == '':
estimated_Run_Time = 90000
    # print("You did not specify how long you wanted to run until so the default value = ", estimated_Run_Time)
steps_TT = int(estimated_Run_Time)
# Initalize Files
edgeLISTa = SP.Edge.create_Edge_Instances()
wb = SP.Network_Period.load_n_create_Excel_NetworkFile(SUMO_outPUT_PREFIX,PERIOD_VARRIABLE,steps_TT,PATH=None)[0]
periodNamesLISTa = SP.Network_Period.load_n_create_Excel_NetworkFile(SUMO_outPUT_PREFIX,PERIOD_VARRIABLE,steps_TT,display=0)[1]

# Take a step(s)
SP.Runner.releaseTraci(Start_Time,typeRun,edgeLISTa,PERIOD_VARRIABLE,SUMO_outPUT_PREFIX,periodNamesLISTa,steps_TT)
SP.Network_Period.fillOutworksheet(SUMO_outPUT_PREFIX,periodCounter = 24,edgeLISTa=edgeLISTa)
traci.close()
# continue01 = 0
# continue01 = str(input("Press x to exit"))
# while continue01 != "x":
    # SP.Runner.releaseTraci(Start_Time,typeRun,edgeLISTa,PERIOD_VARRIABLE,SUMO_outPUT_PREFIX,periodNamesLISTa,steps_TT)
    # continue01 = str(input("Press x to exit"))
    # break
    
### I am up to here ###
# class Edge():
    # # import sumolib
    
    # # edge_list = []
    # def fillOutworksheet(SUMO_outPUT_PREFIX,periodCounter,edgeLISTa):#,periodNamesLISTa):
        # # https://chrisalbon.com/python/data_wrangling/pandas_dataframe_load_xls/
        # BIG CONTRIBUTION PART
        # if periodCounter == 0:
            # return
        # # PATH_Network_DF_Period_0t00_TEMPLATExlsx = '/GitHub/PhD_Modeling/Belmont_AOI_git/Belmont_AOI-runFILES/Network_DF_Period_0t00_TEMPLATE.xlsx'
        # print("\n\n<><><periodCounter = ",int(round(periodCounter)))
        # periodCounter = int(round(periodCounter))#periodCounter = int(round(periodCounter)-1) # "-1" because we want to fill in the sheet from the perivious period.
        # logger_TEMPDF =pd.read_excel(PATH_Network_DF_Period_0t00_TEMPLATExlsx) #creates a template
        # # logger_TEMPDF =pd.read_excel(SP.PATH_Network_DF_Period_0t00_TEMPLATExlsx)
        # counter = 0
        # # for edge_i in Belmont_Ave[:]:
        # for n in range(len(edgeLISTa)): ## I need Belmont_AVEDic[n] == edgeLISTa[n].edgeID == logger_TEMPDF.loc[n,'Belmont_AVEDic_ID']. 1-8-18 saved over tempalte with for loop range(len(edgeLISTa)): newTemplate.loc[i,'Belmont_AVEDic_ID'] = edgeLISTa[i].edgeID 
            # logger_TEMPDF.loc[n,'Total_Vehicles'] = (edgeLISTa[n].truckCount + edgeLISTa[n].carCount)
            # logger_TEMPDF.loc[n,'Total_Trucks'] = edgeLISTa[n].truckCount
            # logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(Belmont_AVEDic[n]),'Total_Trucks'] = edgeLISTa[n].truckCount
            # logger_TEMPDF.loc[n,'ESAL'] = edgeLISTa[n].ESAL_TOT
            # # Condition_RTi = 100.00 - (edgeLISTa[n].ESAL_TOT) * 0.01339
            # # logger_TEMPDF.loc[n,'Condition_Index'] = Condition_RTi
            # maxSpeedo = logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(Belmont_AVEDic[n]),'Original_Max_Speed']
            # # maxSpeedo = edgeLISTa[n].originalMAXSPEED
            # # maxSpeed_i =(maxSpeedo - (maxSpeedo**((100-Condition_RTi)/96))+4.5675) #Units m/s
            # # logger_TEMPDF.loc[n,'Dynamic_Max_Speed'] = maxSpeed_i
            # ##Validating writing below here
            # if n == 4:
                # print("\n[<<[<[<>]>]>>]\nedgeLISTa[4].truckCount = ",edgeLISTa[4].truckCount,"<><>\nlogger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edgeLISTa[n].edgeID),'Total_Trucks']= ",logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edgeLISTa[n].edgeID),'Total_Trucks'],"\n[<<[<[<>]>]>>]\n")
            # # Do I need to reset the logger? Will it slow things down towards the end? 
            # # if len(edgeLISTa[n].vehidLIST) > 100:
                # # if n == 48:
                    # # memoryUse = py.memory_info()[0]/2.**30  # memory use in GB...I think
                    # # print('memory use:', memoryUse)
                    # # edgeLISTa[n]._resetLOGGER(edgeLISTa)
                    # # memoryUse = py.memory_info()[0]/2.**30  # memory use in GB...I think
                    # print('memory use:', memoryUse)
                # edgeLISTa[n]._resetLOGGER(edgeLISTa)
                # Needed for metrics
        # self.meanSpeed = int(traci.edge.getLastStepMeanSpeed(self.edgeID))
        # self.PCR = 0
        # self.IRI = 0
        # self.AVE_QLTH = 0
        # self.ADDT_rand = 0
        # self.ADDT_Calc = 0
        # self.AGE_0 = 0
        # self.SNC = 0
        # self.layer_1_dpth_in = 0
        # self.layer1_6in_35_50 = 0
        # self.layer2_6in_10_25 = 0
        # self.layer3_12in_5_17 = 0
        # self.ASS_CALI = 0
        ## NOW SAVE THIS DATAFRAME [[logger_TEMPDF]] TO THE SHEET THAT IT CAME FROM IN THE WORKBOOKER
        # Network_Period.myWrite_to_excel(logger_TEMPDF,periodCounter,SUMO_outPUT_PREFIX,edgeLISTa)
        ## Change max speed of edge based on roadway damage
        # Edge.setNewMaxSpeed(edgeLISTa,logger_TEMPDF)
        # Edge.getMaxSpeed(edgeLISTa)
        # print("Testing edgeLISTa[48].__dict__ ...\n",edgeLISTa[48].__dict__,"\n")

# SP.traci.edge.adaptTraveltime('387423966',999999)
# SP.traci.edge.setEffort('387423966',9999999)

# SP.traci.edge.getAdaptedTraveltime('387423966',int(SP.traci.simulation.getCurrentTime()/1000))
# SP.traci.edge.getEffort('387423966',int(SP.traci.simulation.getCurrentTime()/1000))

# ##workin on saving excel correctly
# import openpyxl as OPENxlsx
# workBooker = OPENxlsx.Workbook()
# PATH_to_Save_to = "/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES/Network_" +SUMO_outPUT_PREFIX + "_PeriodBook.xlsx"
# workBooker = load_workbook(PATH_to_Save_to)
# workBooker.create_sheet(title=periodNamesLISTa[periodCounter])





    # call each edge
        # get vehID_List
    # call logger
        # compare vehID_List with logger

        # Add new vehIDs to loggers
            # compute stats for vehID
    # Every minute update Network Period Files
        # Update Sumo Edge Parameters
        
        
        
# edge_i_cashe_csv = pd.read_csv(PATH_edge_i_cashe_TEMPLATE)
# edge_cashe_LOAD_PATH = "/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES" + "/" + PC.condition.Belmont_AVEDic[edge_i] +".csv"

###Run until we have values
# while bool(Edge_i_VehIDs_lastStep_j) is False:
     # traci.simulationStep()
     # Runner.thingsTodoWhileStepping(edgeLISTa,PERIOD_VARRIABLE,periodCounter,SUMO_outPUT_PREFIX,periodNamesLISTa)
     # Edge_i_VehIDs_lastStep_j = traci.edge.getLastStepVehicleIDs(edgeLISTa[4].edgeID)
     
     

    
# Try a few things:
# change up dictionary
# hard code
# add hourly cars to evaluate network performance


### Changing up Template 1-8-18 ##

# PATH_Network_DF_Period_0t00_TEMPLATExlsx = '/GitHub/PhD_Modeling/Belmont_AOI_git/Belmont_AOI-runFILES/Network_DF_Period_0t00_TEMPLATE.xlsx'
# newTemplate = pd.read_excel(PATH_Network_DF_Period_0t00_TEMPLATExlsx)
# for i in range(len(edgeLISTa)):
    # newTemplate.loc[i,'Belmont_AVEDic_ID'] = edgeLISTa[i].edgeID
# pd.DateFrame(newTemplate)
# # newTemplate.to_excel(PATH_Network_DF_Period_0t00_TEMPLATExlsx)
# for i in range(75,len(edgeLISTa)):
    # if Belmont_AVEDic[i] == edgeLISTa[i].edgeID == logger_TEMPDF.loc[i,'Belmont_AVEDic_ID']:
        # print(" Matches ",edgeLISTa[i].edgeID," Hell's Yeah!")
        
# edgeLISTa[n].truckCount == logger_TEMPDF.loc[n,'Total_Trucks']

# logger_TEMPDF.loc[n,'Total_Trucks'] == edgeLISTa[n].truckCount
# wsdf.loc[n,7] == edgeLISTa[n].truckCount


# edges = {SP.Belmont_AVEDic[edge]: SP.Edge() for edge in SP.Belmont_Ave}
    # %recall = SP.Edge(str(edge_i))
    # setattr(self, edgeNAME, SP.Edge(str(edge_i)))
    # # exec(edgeNAME + " = SP.Edge(str(edge_i))")
    # # str("edge_"+str(SP.Belmont_AVEDic[edge_i])) = SP.Edge(str(edge_i))
# # edge_73e = SP.Edge('424978646') or now with the set function
# edge_73 = SP.Edge()
# edge_73.set('424978646',73)
    # # for edge_i in SP.Belmont_Ave[:]:
        # # edgeNAME = "edge_"+str(SP.Belmont_AVEDic[edge_i])
        # # print(edgeNAME," - ", str(edge_i))
        # # setattr(SP.Edge,edgeNAME,str(edge_i))
        
# for i in range(10):
    # tabber = "a\t"
    # print(tabber,"a",end='\r',flush=True)
    # tabber = str(tabber+"\t")

## How to tell if periodCounter is a whole number    
# import time
# periodCounter = 0
# PERIOD_VARRIABLE=1800
# for i in range(3600):
# periodCounter += (1/PERIOD_VARRIABLE)
# if round(periodCounter,4).is_integer():# == int(periodCounter):  #% 1 == 0:
    # whole_numb = round(periodCounter,4)
    # print("Whole Number?", whole_numb,"\n")
# else:
    # time.sleep(0.005)
        # print("periodCounter = ",(periodCounter))#,end='\r',flush=True)