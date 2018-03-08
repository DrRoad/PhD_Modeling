
# coding: utf-8

# ##Metric Collector
# #Last worked on 3/7/2018

# In[73]:

import numpy as np
import traci
import traci.constants as tc
import re
import time
import sumolib
# import cProfile, pstats , io
import pandas as pd
import openpyxl as OPENxlsx
pd.set_option('display.max_rows', 5)


# In[60]:

import sumoPython_git_B as SP


# In[ ]:

def collect_metric_data_PAVEMENT():
    pass

def collect_metric_data_DRIVER():
    pass

def collect_metric_data_NETWORK():
    pass


# In[38]:

edge_t0_PATH = '/Dropbox/Phd_R_Ms/PhD_Modeling_DB_GIT/BMAOI_Dbox-DataFrames/Network_BMAOI-RUN_A_January_Monday_WeekTest_11_PeriodBook.xlsx'
edge_t0_DF = pd.read_excel(edge_t0_PATH)
edge_t0_PATH_TEMPLATE = '/Dropbox/Phd_R_Ms/PhD_Modeling_DB_GIT/Belmont_AOI_git/Belmont_AOI-runFILES/Network_DF_Period_0t00_TEMPLATE.xlsx'


# In[78]:

wb =  OPENxlsx.load_workbook(filename = edge_t0_PATH)
wsDICT = dict()
for sheet in range(len(wb.get_sheet_names())):
    wsDICT.update({sheet : wb[wb.get_sheet_names()[sheet]]})
ws = wb[wb.get_sheet_names()[5]]
print("Current Worksheet :-: ",ws)
wsDF = pd.DataFrame(ws.values)
wsDF = wsDF.rename(columns=wsDF.iloc[0]).drop(wsDF.index[0])
wsDF = wsDF.reset_index(drop=True)
wsDICT[5]


# In[89]:

dayAVE_DF = pd.read_excel(edge_t0_PATH_TEMPLATE_METRIC)
dayAVE_DF


# In[92]:

edge_t0_PATH_TEMPLATE_METRIC = '/Dropbox/Phd_R_Ms/PhD_Modeling_DB_GIT/Belmont_AOI_git/Belmont_AOI-runFILES/Network_DF_Period_0t00_TEMPLATE.xlsx'
dayAVE_DF = pd.read_excel(edge_t0_PATH_TEMPLATE_METRIC)
ws = wb[wb.get_sheet_names()[5]]
# for sheet_i in range(5:len(wb.get_sheet_names())):
# ws = wb[wb.get_sheet_names()[sheet_i]]
#### tab over below
print("Current Worksheet :-: ",ws)
wsDF = pd.DataFrame(ws.values)
wsDF = wsDF.rename(columns=wsDF.iloc[0]).drop(wsDF.index[0])
wsDF = wsDF.reset_index(drop=True)
for rd in range(len(wsDF.index)):
    print(wsDF.loc[rd,'Belmont_AVEDic_ID'])
    


# In[75]:




# In[79]:

wsDF


# In[61]:

logger_TEMPDF = wsDF

edgeLISTa = SP.Edge.create_Edge_Instances(wsDF)


# In[77]:

logger_TEMPDF = wsDF
logger_TEMPDF


# In[81]:

# edgeLISTa[rd].ESAL_TOT


# In[83]:

# segments from Network_Period.fillOutworksheet(wsDF,SUMO_outPUT_PREFIX,periodCounter,edgeLISTa)
edge_t0_PATH_SET = '/Dropbox/Phd_R_Ms/PhD_Modeling_DB_GIT/BMAOI_Dbox-DataFrames/Network_BMAOI-RUN_A_January_Monday_WeekTest_11_PeriodBook.xlsx'
def set_edgeLISTa_with(edge_t0_PATH_SET):
    wb =  OPENxlsx.load_workbook(filename = edge_t0_PATH_SET)
    wsDICT = dict()
    for sheet in range(len(wb.get_sheet_names())):
        wsDICT.update({sheet : wb[wb.get_sheet_names()[sheet]]})
    ws = wb[wb.get_sheet_names()[5]]
    print("Current Worksheet :-: ",ws)
    wsDF = pd.DataFrame(ws.values)
    wsDF = wsDF.rename(columns=wsDF.iloc[0]).drop(wsDF.index[0])
    wsDF = wsDF.reset_index(drop=True)
    for rd in range(len(edgeLISTa)): ## I need Belmont_AVEDic[rd] == edgeLISTa[rd].edgeID == wsDF.loc[rd,'Belmont_AVEDic_ID']. 1-8-18 saved over tempalte with for loop range(len(edgeLISTa)): newTemplate.loc[i,'Belmont_AVEDic_ID'] = edgeLISTa[i].edgeID 
        edgeLISTa[rd].carCount = wsDF.loc[rd,'Total_Vehicles'] - wsDF.loc[rd,'Total_Trucks']
        edgeLISTa[rd].truckCount = wsDF.loc[rd,'Total_Trucks']
        ## wsDF.loc[wsDF['Belmont_AVEDic_ID'].str.contains(Belmont_AVEDic[rd]),'Total_Trucks'] = edgeLISTa[rd].truckCount
        edgeLISTa[rd].ESAL_TOT = wsDF.loc[rd,'ESAL_TOT']
        Condition_RTi = 100.00 - (wsDF.loc[0,'ESAL_TOT']) * 0.01339
        edgeLISTa[rd].Condition_Index = Condition_RTi
#         edgeLISTa[rd].ADDT_Calc = wsDF.loc[rd,'AADT_rand']
        # maxSpeedo = wsDF.loc[wsDF['Belmont_AVEDic_ID'].str.contains(Belmont_AVEDic[rd]),'Original_Max_Speed']
        maxSpeedo = edgeLISTa[rd].originalMAXSPEED
        maxSpeed_i =(maxSpeedo - (maxSpeedo**((100-Condition_RTi)/96))+4.5675) #Units m/s
        wsDF.loc[rd,'Dynamic_Max_Speed'] = maxSpeed_i
        edgeLISTa[rd].Dynamic_Max_Speed = maxSpeed_i
    ## Needed for metrics
        edgeLISTa[rd].meanSpeed = 25#int(traci.edge.getLastStepMeanSpeed(edgeLISTa[rd].edgeID))
        edgeLISTa[rd].PCR = 90-0.6349 * (np.exp((edgeLISTa[rd].AGE_0+edgeLISTa[rd].AGE_t/31556926)**0.4203)-1) * np.log(edgeLISTa[rd].ESAL_TOT/((edgeLISTa[rd].SNC)**2.7062))
        edgeLISTa[rd].IRI = 52 + 8.1 * ((int(edgeLISTa[rd].AGE_t)/31556926))+0.0009*edgeLISTa[rd].ADDT_Calc
        edgeLISTa[rd].AVE_QLTH = 0
    #     edgeLISTa[rd].OCPNY = traci.edge.getLastStepOccupancy(edgeLISTa[rd].edgeID)
set_edgeLISTa_with(edge_t0_PATH_SET)


# In[87]:

edgeLISTa[1].__dict__


# In[72]:

# segments from Network_Period.fillOutworksheet(wsDF,SUMO_outPUT_PREFIX,periodCounter,edgeLISTa)
for rd in range(len(edgeLISTa)): ## I need Belmont_AVEDic[rd] == edgeLISTa[rd].edgeID == logger_TEMPDF.loc[rd,'Belmont_AVEDic_ID']. 1-8-18 saved over tempalte with for loop range(len(edgeLISTa)): newTemplate.loc[i,'Belmont_AVEDic_ID'] = edgeLISTa[i].edgeID 
    logger_TEMPDF.loc[rd,'Total_Vehicles'] = (edgeLISTa[rd].truckCount + edgeLISTa[rd].carCount)
    logger_TEMPDF.loc[rd,'Total_Trucks'] = edgeLISTa[rd].truckCount
    ## logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(Belmont_AVEDic[rd]),'Total_Trucks'] = edgeLISTa[rd].truckCount
    logger_TEMPDF.loc[rd,'ESAL_TOT'] = edgeLISTa[rd].ESAL_TOT
    Condition_RTi = 100.00 - (edgeLISTa[rd].ESAL_TOT) * 0.01339
    logger_TEMPDF.loc[rd,'Condition_Index'] = Condition_RTi
    # maxSpeedo = logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(Belmont_AVEDic[rd]),'Original_Max_Speed']
    maxSpeedo = edgeLISTa[rd].originalMAXSPEED
    maxSpeed_i =(maxSpeedo - (maxSpeedo**((100-Condition_RTi)/96))+4.5675) #Units m/s
    logger_TEMPDF.loc[rd,'Dynamic_Max_Speed'] = maxSpeed_i
    edgeLISTa[rd].Dynamic_Max_Speed = maxSpeed_i
## Needed for metrics
    edgeLISTa[rd].meanSpeed = 25#int(traci.edge.getLastStepMeanSpeed(edgeLISTa[rd].edgeID))
    edgeLISTa[rd].PCR = 90-0.6349 * (np.exp((edgeLISTa[rd].AGE_0+edgeLISTa[rd].AGE_t/31556926)**0.4203)-1) * np.log(edgeLISTa[rd].ESAL_TOT/((edgeLISTa[rd].SNC)**2.7062))
    edgeLISTa[rd].IRI = 52 + 8.1 * ((int(edgeLISTa[rd].AGE_t)/31556926))+0.0009*edgeLISTa[rd].ADDT_Calc
    edgeLISTa[rd].AVE_QLTH = 0
#     edgeLISTa[rd].OCPNY = traci.edge.getLastStepOccupancy(edgeLISTa[rd].edgeID)


# In[34]:




# In[29]:

wsDF


# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:

def fillOutworksheet(edge_t0_DF,SUMO_outPUT_PREFIX,periodCounter,edgeLISTa):#,periodNamesLISTa):
    #https://chrisalbon.com/python/data_wrangling/pandas_dataframe_load_xls/
    #### BIG CONTRIBUTION PART
    print("\n\n\t\t\tFilling out worksheet")
    if periodCounter == 0:
        return
    # PATH_Network_DF_Period_0t00_TEMPLATExlsx = '/Dropbox/Phd_R_Ms/PhD_Modeling_DB_GIT/Belmont_AOI_git/Belmont_AOI-runFILES/Network_DF_Period_0t00_TEMPLATE.xlsx'
    print("\n\n<><><periodCounter = ",int(round(periodCounter)))
    periodCounter = int(round(periodCounter))#periodCounter = int(round(periodCounter)-1) # "-1" because we want to fill in the sheet from the perivious period.
    logger_TEMPDF = edge_t0_DF#pd.read_excel(PATH_Network_DF_Period_0t00_TEMPLATExlsx) #creates a template
    # logger_TEMPDF =pd.read_excel(SP.PATH_Network_DF_Period_0t00_TEMPLATExlsx)
    counter = 0
    #for edge_i in Belmont_Ave[:]:
    for rd in range(len(edgeLISTa)): ## I need Belmont_AVEDic[rd] == edgeLISTa[rd].edgeID == logger_TEMPDF.loc[rd,'Belmont_AVEDic_ID']. 1-8-18 saved over tempalte with for loop range(len(edgeLISTa)): newTemplate.loc[i,'Belmont_AVEDic_ID'] = edgeLISTa[i].edgeID 
        logger_TEMPDF.loc[rd,'Total_Vehicles'] = (edgeLISTa[rd].truckCount + edgeLISTa[rd].carCount)
        logger_TEMPDF.loc[rd,'Total_Trucks'] = edgeLISTa[rd].truckCount
        ## logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(Belmont_AVEDic[rd]),'Total_Trucks'] = edgeLISTa[rd].truckCount
        logger_TEMPDF.loc[rd,'ESAL_TOT'] = edgeLISTa[rd].ESAL_TOT
        Condition_RTi = 100.00 - (edgeLISTa[rd].ESAL_TOT) * 0.01339
        logger_TEMPDF.loc[rd,'Condition_Index'] = Condition_RTi
        # maxSpeedo = logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(Belmont_AVEDic[rd]),'Original_Max_Speed']
        maxSpeedo = edgeLISTa[rd].originalMAXSPEED
        maxSpeed_i =(maxSpeedo - (maxSpeedo**((100-Condition_RTi)/96))+4.5675) #Units m/s
        logger_TEMPDF.loc[rd,'Dynamic_Max_Speed'] = maxSpeed_i
        edgeLISTa[rd].Dynamic_Max_Speed = maxSpeed_i
    ## Needed for metrics
        edgeLISTa[rd].meanSpeed = int(traci.edge.getLastStepMeanSpeed(edgeLISTa[rd].edgeID))
        edgeLISTa[rd].PCR = 90-0.6349 * (np.exp((edgeLISTa[rd].AGE_0+edgeLISTa[rd].AGE_t/31556926)**0.4203)-1) * np.log(edgeLISTa[rd].ESAL_TOT/((edgeLISTa[rd].SNC)**2.7062))
        edgeLISTa[rd].IRI = 52 + 8.1 * ((int(edgeLISTa[rd].AGE_t)/31556926))+0.0009*edgeLISTa[rd].ADDT_Calc
        edgeLISTa[rd].AVE_QLTH = 0
        edgeLISTa[rd].OCPNY = traci.edge.getLastStepOccupancy(edgeLISTa[rd].edgeID)
        # ##Validating writing below here
        # if n == 4:
            # print("\n[<<[<[<>]>]>>]\nedgeLISTa[4].truckCount = ",edgeLISTa[4].truckCount,"<><>\nlogger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edgeLISTa[rd].edgeID),'Total_Trucks']= ",logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edgeLISTa[rd].edgeID),'Total_Trucks'],"\n[<<[<[<>]>]>>]\n")
        ## Do I need to reset the logger? Will it slow things down towards the end? 
        if len(edgeLISTa[rd].vehidLIST) > 100:
            # if n == 48:
            memoryUse = py.memory_info()[0]/2.**30  # memory use in GB...I think
            print('memory use:', memoryUse)
            edgeLISTa[rd]._resetLOGGER(edgeLISTa)
            memoryUse = py.memory_info()[0]/2.**30  # memory use in GB...I think
            print('memory use:', memoryUse)
            edgeLISTa[rd]._resetLOGGER(edgeLISTa)

    ### NOW SAVE THIS DATAFRAME [[logger_TEMPDF]] TO THE SHEET THAT IT CAME FROM IN THE WORKBOOKER
    wsdf = Network_Period.myWrite_to_excel(logger_TEMPDF,periodCounter,SUMO_outPUT_PREFIX,edgeLISTa)[1]
    ### Change max speed of edge based on roadway damage
    Edge.setNewMaxSpeed(edgeLISTa,logger_TEMPDF)
    Edge.getMaxSpeed(edgeLISTa,logger_TEMPDF)
    print("Testing edgeLISTa[48].__dict__ ...\n",edgeLISTa[48].__dict__,"\n======",SUMO_outPUT_PREFIX,"======")
    print("\t\t\t\t======",str(SUMO_outPUT_PREFIX),"======")
    return wsdf#, wb 
        
def myWrite_to_excel(logger_TEMPDF,periodCounter,SUMO_outPUT_PREFIX,edgeLISTa):
    # https://openpyxl.readthedocs.io/en/default/tutorial.html#data-storage   
    # http://openpyxl.readthedocs.io/en/default/pandas.html#working-with-pandas-dataframes
    ### TRY THIS http://pbpython.com/improve-pandas-excel-output.html and then this https://codereview.stackexchange.com/questions/180405/writing-excelsheet-from-python-dataframe
    # from openpyxl import load_workbook
    ## openpyxl.worksheet.Worksheet.cell() method from  
    PATH_to_Save_to = "/Dropbox/Phd_R_Ms/PhD_Modeling_DB_GIT/BMAOI_Dbox-DataFrames/Network_" +str(SUMO_outPUT_PREFIX) + "_PeriodBook.xlsx"
    OLD_PATH_to_Save_to = "/Sumo/runs/BelmontC_AOI_main/BelmontC_AOI-outPUT/BMAOI_C-DataFrames/BMAOI_edgeCasheFILES/Network_" +str(SUMO_outPUT_PREFIX) + "_PeriodBook.xlsx"
    wb =  OPENxlsx.load_workbook(filename = PATH_to_Save_to)
    periodCounter = int(round(periodCounter))
    ws = wb[wb.get_sheet_names()[periodCounter-1]]
    print("\n[[[<><><>]]]\nStarting myWrite_to_excel()\n\t PATH_to_Save_to: ",PATH_to_Save_to,"\n\t Current Worksheet: ",ws,"Current periodCounter = ",periodCounter,"\n\n[[[<><><>]]]")
    wsROW = 0
    for i in range(logger_TEMPDF.shape[0]):
        wsROW +=1 
        wsCOL =0
        for j in range(logger_TEMPDF.shape[1]):
            wsCOL +=1
            if i == 1:
                ws.cell(row=wsROW, column=wsCOL, value=logger_TEMPDF.columns[j])
                # ws.cell(row=wsROW, column=wsCOL, value=logger_TEMPDF.columns[j])
                if i == 1 and j == 1:
                    print("\t\t\t\t<Writing Headers>")
                ws.cell(row=wsROW+1, column=wsCOL, value=logger_TEMPDF.iloc[i,j])
            else:
                ws.cell(row=wsROW+1, column=wsCOL, value=logger_TEMPDF.iloc[i,j])
        if i == 7:
            wsdf = pd.DataFrame(ws.values,columns=logger_TEMPDF.columns)
            print("\nPlease be true ... ", wsdf.iloc[6,7] == edgeLISTa[6].truckCount)
            # logger_TEMPDF.loc[logger_TEMPDF['Belmont_AVEDic_ID'].str.contains(edgeLISTa[6].edgeID),'Total_Trucks']
    wsdf = logger_TEMPDF#pd.DataFrame(ws.values,columns=logger_TEMPDF.columns)
    wb.save(PATH_to_Save_to)
    print("\n>>>Data written to:: wb[wb.get_sheet_names()[periodCounter-1]]= ",wb[wb.get_sheet_names()[periodCounter-1]],"\n\t\t\t\t\t this should match ws: ",ws,"\n[[[[<<<<Next Period>>>>]]]]: ",wb.get_sheet_names()[periodCounter])#,"\n\wsdf = ",wsdf)
    return wb, wsdf

